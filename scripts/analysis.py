#!/usr/bin/env python3
"""
analysis.py — single source of truth for every number reported in the ETCG manuscript.

Round 1 reviewers found that the same quantities were reported with different values in
Table 4, Section 5.1, and the Threats to Validity section. Root cause: the study was
upgraded from two conditions to three late in preparation; tables and abstract were
hand-updated, prose and figures were not.

This script exists so that cannot happen again. It reads exactly one scoring file, computes
every reported quantity once, and emits:

    results/results.json      every value, keyed by manuscript label
    results/macros.tex        a \\newcommand for every value, for use in prose
    results/table-*.tex       generated LaTeX tables, for \\input into main.tex
    results/TRACEABILITY.md   the appendix mapping every value to its computation
    figures/*.pdf             regenerated figures (via figures.py)

No number may be typed by hand into main.tex. Prose cites \\ETCG... macros; tables are
\\input. The --audit-tex mode flags any decimal literal in the manuscript that this script
did not produce.

Usage:
    python analysis.py
    python analysis.py --no-figures
    python analysis.py --audit-tex ../../paper-2-EMSE/submission/main.tex

Author: Arbaz Surti
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
from scipy import stats

# ── Configuration ────────────────────────────────────────────────────────────

CONDITIONS = ["etcg", "intermediate", "baseline"]
CONDITION_LABELS = {"etcg": "ETCG", "intermediate": "Intermediate", "baseline": "Baseline"}
DIMENSIONS = ["specificity", "testability", "risk_coverage", "clarity", "actionability"]
DIMENSION_LABELS = {
    "specificity": "Specificity",
    "testability": "Testability",
    "risk_coverage": "Risk Coverage",
    "clarity": "Clarity",
    "actionability": "Actionability",
}
# The three pairwise contrasts, named for what each isolates.
CONTRASTS = [
    ("etcg", "baseline", "full_framework", "ETCG versus Baseline (full framework)"),
    ("intermediate", "baseline", "role_guidance", "Intermediate versus Baseline (role + guidance)"),
    ("etcg", "intermediate", "schema", "ETCG versus Intermediate (output schema)"),
]

# Specification richness groups. SPEC-01..18 are sparse (description-only, reused from the
# ATSR study); SPEC-19..25 conform fully to the ETCG input model.
SPARSE_SPECS = {f"SPEC-{i:02d}" for i in range(1, 19)}
STRUCTURED_SPECS = {f"SPEC-{i:02d}" for i in range(19, 26)}

# Specification-level means are averages of five scores on a coarse 15-point scale, so
# genuine ties in |difference| are common. Whether such a tie lands as exactly equal or
# 1e-16 apart depends on floating-point summation order, and Wilcoxon's tie correction is
# sensitive to that: the same test on the same data returned p = 0.039888 or 0.039905
# depending only on the order rows arrived in. Rounding the means to a precision far below
# any reported digit makes ties exact, which is both the arithmetically correct treatment
# and a precondition for a replication package that returns the same number every run.
SPEC_MEAN_DP = 10

# Baseline cells whose charter numbering changed when the preamble-split defect was
# repaired (see scripts/repair_baseline_split.py). Charter ids in these cells do not
# denote the same charter before and after the repair.
REPAIRED_SPECS = ["SPEC-01", "SPEC-06", "SPEC-10", "SPEC-23"]
REPAIRED_CELLS = {(s, "baseline", f"BL-{i:02d}")
                  for s in REPAIRED_SPECS for i in range(1, 6)}

# The scoring file exactly as it stood before repair_baseline_split.py ran. Retained as a
# required package input so the manuscript's disclosure of what that repair changed (the
# provenance paragraph in Threats to Validity) can be recomputed rather than taken on
# trust: prerepair_headline() reads it and register_all() emits the superseded contrasts
# next to the corrected ones. Do not delete when trimming the other *.pre-splitfix-backup
# files -- this one is consumed by the analysis.
PREREPAIR_SCORES_NAME = "etcg-scores.json.pre-splitfix-backup"

SCALE_MIN, SCALE_MAX = 1, 3
MAX_CHARTER_POINTS = len(DIMENSIONS) * SCALE_MAX
CHARTERS_PER_CELL = 5


# ── Registry: every reported value passes through here ───────────────────────


class Registry:
    """Records every computed value with its description and how it was derived.

    This is the mechanism that makes the manuscript traceable: a value that is not
    registered cannot appear in a table, a macro, or the traceability appendix.
    """

    def __init__(self) -> None:
        self._entries: dict[str, dict] = {}

    def put(self, key: str, value, *, desc: str, fmt: str = "{:.2f}", method: str = "") -> None:
        if key in self._entries:
            raise KeyError(f"duplicate registry key: {key!r}")
        if isinstance(value, (np.floating, np.integer)):
            value = value.item()
        self._entries[key] = {
            "value": value,
            "desc": desc,
            "fmt": fmt,
            "method": method,
        }

    def get(self, key: str):
        return self._entries[key]["value"]

    def rendered(self, key: str) -> str:
        e = self._entries[key]
        return render(e["value"], e["fmt"])

    def keys(self) -> list[str]:
        return list(self._entries)

    def as_dict(self) -> dict:
        return {
            k: {"value": e["value"], "description": e["desc"], "method": e["method"],
                "formatted": render(e["value"], e["fmt"])}
            for k, e in self._entries.items()
        }

    def entries(self):
        return self._entries.items()


def render(value, fmt: str) -> str:
    """Format a value for the manuscript.

    A bare ASCII hyphen sets as a hyphen, not a minus sign, so every negative value is
    emitted as $-$ in LaTeX. The 'diff' formats additionally carry an explicit $+$ so a
    difference always reads with its direction visible.
    """
    if value is None:
        return "--"
    if fmt == "p":
        return format_p(value)
    if fmt == "int":
        return f"{int(round(value)):d}"
    if fmt == "pct0":
        return f"{value:.0f}"

    signed = fmt.startswith("diff")
    if signed:
        places = int(fmt[4:]) if len(fmt) > 4 else 2
        body = f"{abs(value):.{places}f}"
        return f"$-${body}" if value < 0 else f"$+${body}"

    if isinstance(value, float) and value < 0:
        return "$-$" + fmt.format(abs(value))
    return fmt.format(value)


def format_p(p: float) -> str:
    """Manuscript p-value convention: three decimals, or < 0.001 below that."""
    if p < 0.001:
        return "$<$0.001"
    return f"{p:.3f}"


DIGIT_WORDS = {"0": "Zero", "1": "One", "2": "Two", "3": "Three", "4": "Four",
               "5": "Five", "6": "Six", "7": "Seven", "8": "Eight", "9": "Nine"}


def key_to_macro(key: str) -> str:
    r"""Map a registry key to a LaTeX command name.

    LaTeX command names admit only letters, so digits are spelled out and separators
    dropped: 'rq1.etcg.mean' -> '\ETCGRqOneEtcgMean'.
    """
    parts = re.split(r"[._\-]+", key)
    out = []
    for part in parts:
        chunk = "".join(DIGIT_WORDS[c] if c.isdigit() else c for c in part)
        out.append(chunk[:1].upper() + chunk[1:])
    name = "ETCG" + "".join(out)
    if not name.isalpha():
        raise ValueError(f"key {key!r} produced non-alphabetic macro {name!r}")
    return name


# ── Data loading and design validation ───────────────────────────────────────


def load_scores(path: Path) -> tuple[list[dict], dict]:
    with path.open(encoding="utf-8") as fh:
        blob = json.load(fh)
    scores = blob["scores"]
    meta = blob.get("run_metadata", {})
    return scores, meta


def validate_design(scores: list[dict]) -> dict:
    """Fail loudly if the corpus is not the fully-crossed design the analysis assumes.

    Reviewer 2's pseudoreplication comment turns on the design being repeated-measures.
    If that ever stops being true, every paired test below becomes invalid, so this is
    checked rather than assumed.
    """
    problems = []

    errored = [r for r in scores if "error" in r]
    if errored:
        problems.append(f"{len(errored)} scoring records carry an error field")

    cells = Counter((r["spec_id"], r["condition"]) for r in scores)
    specs = sorted({r["spec_id"] for r in scores})
    conds = sorted({r["condition"] for r in scores})

    if set(conds) != set(CONDITIONS):
        problems.append(f"conditions {conds} != expected {CONDITIONS}")

    expected_cells = len(specs) * len(CONDITIONS)
    if len(cells) != expected_cells:
        problems.append(f"{len(cells)} spec x condition cells, expected {expected_cells}")

    bad = {k: v for k, v in cells.items() if v != CHARTERS_PER_CELL}
    if bad:
        problems.append(f"{len(bad)} cells do not hold {CHARTERS_PER_CELL} charters: {list(bad)[:5]}")

    known = SPARSE_SPECS | STRUCTURED_SPECS
    unknown = set(specs) - known
    if unknown:
        problems.append(f"specs not assigned to a richness group: {sorted(unknown)}")

    for r in scores:
        for d in DIMENSIONS:
            v = r["scores"].get(d)
            if v is None or not (SCALE_MIN <= v <= SCALE_MAX):
                problems.append(f"{r['spec_id']}/{r['charter_id']}: {d} = {v!r} out of scale")
                break

    if problems:
        raise SystemExit("DESIGN VALIDATION FAILED:\n  - " + "\n  - ".join(problems))

    return {"n_specs": len(specs), "n_conditions": len(conds), "n_charters": len(scores),
            "specs": specs}


# ── Aggregation ──────────────────────────────────────────────────────────────


def build_matrices(scores: list[dict], specs: list[str]) -> dict:
    """Aggregate to the specification level — the unit of analysis.

    Charters generated from the same specification share an input context and a generation
    call, so they are not independent observations. Every inferential test below runs on
    these per-specification means, with specification as the repeated-measures unit.
    """
    idx = {s: i for i, s in enumerate(specs)}
    n = len(specs)

    charter_pct = {c: [] for c in CONDITIONS}
    spec_pct = {c: np.full(n, np.nan) for c in CONDITIONS}
    spec_dim = {c: {d: np.full(n, np.nan) for d in DIMENSIONS} for c in CONDITIONS}
    charter_dim = {c: {d: [] for d in DIMENSIONS} for c in CONDITIONS}

    grouped = defaultdict(list)
    for r in scores:
        grouped[(r["spec_id"], r["condition"])].append(r)

    for (spec, cond), rows in grouped.items():
        i = idx[spec]
        pcts = [r["percentage"] for r in rows]
        spec_pct[cond][i] = round(float(np.mean(pcts)), SPEC_MEAN_DP)
        charter_pct[cond].extend(pcts)
        for d in DIMENSIONS:
            vals = [r["scores"][d] for r in rows]
            spec_dim[cond][d][i] = round(float(np.mean(vals)), SPEC_MEAN_DP)
            charter_dim[cond][d].extend(vals)

    for c in CONDITIONS:
        assert not np.isnan(spec_pct[c]).any(), f"missing spec-level data for {c}"

    return {
        "specs": specs,
        "spec_pct": spec_pct,
        "spec_dim": spec_dim,
        "charter_pct": {c: np.array(v, dtype=float) for c, v in charter_pct.items()},
        "charter_dim": {c: {d: np.array(v, dtype=float) for d, v in dd.items()}
                        for c, dd in charter_dim.items()},
    }


# ── Statistics ───────────────────────────────────────────────────────────────


def ci_mean(x: np.ndarray, conf: float = 0.95) -> tuple[float, float]:
    """t-based confidence interval on a mean."""
    x = np.asarray(x, dtype=float)
    n = len(x)
    if n < 2:
        return (float("nan"), float("nan"))
    se = x.std(ddof=1) / math.sqrt(n)
    t = stats.t.ppf(0.5 + conf / 2, df=n - 1)
    m = float(x.mean())
    return (m - t * se, m + t * se)


def bootstrap_ci_mean(x: np.ndarray, conf: float = 0.95, *, n_resamples: int = 10_000,
                      seed: int = 0) -> tuple[float, float]:
    """Deterministic percentile-bootstrap confidence interval for a sample mean.

    A t interval can extend past the 0--100 bounds of a percentage score. Resampling the
    observed specification means instead keeps the plotted interval within the scale and
    is more appropriate for the small, ceilinged RQ4 groups. A fixed seed makes the
    generated figure and traceability values reproducible.
    """
    x = np.asarray(x, dtype=float)
    if len(x) < 2:
        return (float("nan"), float("nan"))
    rng = np.random.default_rng(seed)
    draws = x[rng.integers(0, len(x), size=(n_resamples, len(x)))].mean(axis=1)
    alpha = (1.0 - conf) / 2.0
    return (float(np.quantile(draws, alpha)), float(np.quantile(draws, 1.0 - alpha)))


def rank_biserial_paired(diff: np.ndarray) -> float:
    """Matched-pairs rank-biserial correlation — the effect size that pairs with the
    Wilcoxon signed-rank test. Ranges -1..1; sign follows the direction of the difference."""
    d = np.asarray(diff, dtype=float)
    nz = d[d != 0]
    if nz.size == 0:
        return 0.0
    ranks = stats.rankdata(np.abs(nz))
    w_pos = ranks[nz > 0].sum()
    w_neg = ranks[nz < 0].sum()
    total = w_pos + w_neg
    return float((w_pos - w_neg) / total) if total else 0.0


def paired_contrast(a: np.ndarray, b: np.ndarray, label: str) -> dict:
    """The primary inferential test: paired, at the specification level.

    Wilcoxon signed-rank is primary (the spec-level means are not normally distributed
    and n is small); the paired t-test is reported alongside as a robustness check.
    """
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    diff = a - b
    n = len(diff)
    n_zero = int((diff == 0).sum())

    if np.all(diff == 0):
        w_stat, p_signed = float("nan"), 1.0
    else:
        w_stat, p_signed = stats.wilcoxon(a, b, zero_method="wilcox", method="auto")

    t_stat, p_t = stats.ttest_rel(a, b)
    sd_diff = diff.std(ddof=1)
    dz = float(diff.mean() / sd_diff) if sd_diff > 0 else float("nan")
    lo, hi = ci_mean(diff)

    return {
        "label": label,
        "n_pairs": n,
        "n_zero_diff": n_zero,
        "mean_diff": float(diff.mean()),
        "ci_low": lo,
        "ci_high": hi,
        "w_stat": float(w_stat),
        "p_signed_rank": float(p_signed),
        "t_stat": float(t_stat),
        "p_paired_t": float(p_t),
        "dz": dz,
        "rank_biserial": rank_biserial_paired(diff),
    }


def superseded_unpaired(a: np.ndarray, b: np.ndarray) -> dict:
    """The charter-level Mann-Whitney test the round 1 submission relied on, recomputed
    here on the current (repaired) data.

    Retained and reported ONLY so the point-by-point response can show exactly what the
    corrected analysis changed. It treats 125 clustered charters as independent
    observations and is not a valid test of these data. The value is not the number
    printed in the round 1 submission: it is that submission's *method* re-run on the
    corrected scoring file, so the "before" and "after" differ only in the estimator.
    """
    u, p = stats.mannwhitneyu(a, b, alternative="two-sided")
    return {"u_stat": float(u), "p_value": float(p),
            "note": "SUPERSEDED — pseudoreplicated; charters within a specification are not independent"}


def prerepair_headline(path: Path) -> dict | None:
    """RQ1 condition means and pairwise contrasts on the pre-repair scoring file.

    repair_baseline_split.py corrected a splitter defect that had scored a non-charter
    preamble or heading as a charter in four baseline cells (REPAIRED_SPECS), depressing the baseline
    mean on exactly the ETCG-versus-baseline contrast the earlier analysis called
    significant. The manuscript discloses what the repair changed; this recomputes the
    superseded numbers from the retained pre-repair file so that disclosure is traceable
    rather than asserted. Estimator is identical to the primary analysis: Wilcoxon
    signed-rank on 25 paired per-specification means.
    """
    if not path.exists():
        return None
    scores, _ = load_scores(path)
    specs = sorted({r["spec_id"] for r in scores})
    m = build_matrices(scores, specs)
    out: dict = {"n_specs": len(specs), "conditions": {}, "contrasts": {}}
    for cond in CONDITIONS:
        out["conditions"][cond] = {
            "mean": float(m["spec_pct"][cond].mean()),
            "sd_charter": float(m["charter_pct"][cond].std(ddof=1)),
        }
    for a, b, slug, label in CONTRASTS:
        res = paired_contrast(m["spec_pct"][a], m["spec_pct"][b], label)
        out["contrasts"][slug] = {
            "mean_diff": res["mean_diff"],
            "p_signed_rank": res["p_signed_rank"],
            "ci_low": res["ci_low"],
            "ci_high": res["ci_high"],
        }
    return out


def holm_bonferroni(pvals: dict[str, float]) -> dict[str, float]:
    """Holm-Bonferroni step-down adjustment across a family of tests."""
    items = sorted(pvals.items(), key=lambda kv: kv[1])
    m = len(items)
    adjusted, running = {}, 0.0
    for i, (key, p) in enumerate(items):
        val = min(1.0, (m - i) * p)
        running = max(running, val)  # enforce monotonicity
        adjusted[key] = running
    return adjusted


def mixed_effects(scores: list[dict]) -> dict | None:
    """Linear mixed-effects model: percentage ~ condition, random intercept per specification.

    This is the model Reviewer 2 asked for, and it is fitted on all 375 individual charter
    scores rather than on aggregated means: it models the clustering by specification
    explicitly instead of averaging it away, which is the point of using a mixed model at
    all. The intraclass correlation it reports quantifies how much clustering there actually
    was — that is, how much the round 1 analysis was inflating its effective sample size.

    Fitted twice with different reference levels so all three pairwise contrasts are
    available; the fixed-effect estimates are identical either way, only the contrasts
    reported change.
    """
    try:
        import pandas as pd
        import statsmodels.formula.api as smf
    except ImportError:
        return None

    df = pd.DataFrame([
        {"spec": r["spec_id"], "condition": str(r["condition"]), "pct": float(r["percentage"])}
        for r in scores
    ])

    def fit_with(reference: str):
        formula = f"pct ~ C(condition, Treatment(reference='{reference}'))"
        model = smf.mixedlm(formula, df, groups=df["spec"])
        # lbfgs hits a singular Hessian on these data; Powell and CG both converge.
        last = None
        for method in ("powell", "cg", "bfgs", "lbfgs"):
            try:
                fit = model.fit(reml=True, method=method)
                if fit.converged:
                    return fit, method
                last = f"{method} did not converge"
            except Exception as exc:
                last = f"{method}: {type(exc).__name__}"
        raise RuntimeError(last or "no optimizer converged")

    try:
        fit_base, method = fit_with("baseline")
        fit_inter, _ = fit_with("intermediate")
    except Exception as exc:  # pragma: no cover - reported, not fatal
        return {"error": f"{type(exc).__name__}: {exc}"}

    def term_for(fit, level: str) -> dict | None:
        for name in fit.params.index:
            if name.startswith("Group"):
                continue
            if name.endswith(f"T.{level}]"):
                return {"coef": float(fit.params[name]), "se": float(fit.bse[name]),
                        "z": float(fit.tvalues[name]), "p": float(fit.pvalues[name])}
        return None

    contrasts = {
        "full_framework": term_for(fit_base, "etcg"),
        "role_guidance": term_for(fit_base, "intermediate"),
        "schema": term_for(fit_inter, "etcg"),
    }
    if any(v is None for v in contrasts.values()):
        return {"error": "could not extract all three contrasts from the fitted models"}

    group_var = float(fit_base.cov_re.iloc[0, 0]) if fit_base.cov_re.size else float("nan")
    resid_var = float(fit_base.scale)

    return {
        "formula": "pct ~ condition + (1 | specification)",
        "optimizer": method,
        "reference_level": "baseline",
        "intercept": float(fit_base.params["Intercept"]),
        "contrasts": contrasts,
        "group_var": group_var,
        "resid_var": resid_var,
        "icc": group_var / (group_var + resid_var) if (group_var + resid_var) > 0 else float("nan"),
        "n_obs": int(fit_base.nobs),
        "n_groups": int(df["spec"].nunique()),
        "converged": True,
    }


# ── Scorer stability across runs ─────────────────────────────────────────────


def scorer_stability(scores: list[dict], prior_path: Path) -> dict | None:
    """Compare the current scoring run against an earlier run of the same scorer.

    The ETCG and Baseline conditions were scored twice by GPT-4o at temperature 0, on
    24 and 25 March 2026. That accidental replication is direct evidence on how stable the
    automated evaluator actually is — the manuscript asserted aggregate stability but had
    never measured per-item agreement.

    The earlier file is used ONLY here. It predates the Intermediate condition and must
    never be a source for reported results.

    It also predates the baseline split repair. In the three specifications whose split
    changed, a given charter_id no longer denotes the same charter in both files, so
    joining on the id would compare a charter against whatever previously held its
    number and report the difference as scorer drift. Those cells are excluded: this
    statistic measures how stable the scorer is, not how much the corpus was corrected.
    """
    if not prior_path.exists():
        return None
    try:
        with prior_path.open(encoding="utf-8") as fh:
            prior_rows = json.load(fh)["scores"]
    except (json.JSONDecodeError, KeyError, OSError):
        return None

    def index(rows):
        return {(r["spec_id"], r["condition"], r["charter_id"]): r for r in rows}

    now, before = index(scores), index(prior_rows)
    common = sorted((set(now) & set(before)) - REPAIRED_CELLS)
    if not common:
        return None

    deltas = np.array([now[k]["percentage"] - before[k]["percentage"] for k in common])
    flips = sum(1 for k in common for d in DIMENSIONS
                if now[k]["scores"][d] != before[k]["scores"][d])
    n_ratings = len(common) * len(DIMENSIONS)

    per_condition = {}
    for cond in sorted({k[1] for k in common}):
        keys = [k for k in common if k[1] == cond]
        a = np.array([now[k]["percentage"] for k in keys])
        b = np.array([before[k]["percentage"] for k in keys])
        per_condition[cond] = {
            "n": len(keys),
            "mean_before": float(b.mean()),
            "mean_after": float(a.mean()),
            "mean_shift": float(a.mean() - b.mean()),
            "n_changed": int((a != b).sum()),
        }

    return {
        "prior_file": str(prior_path),
        "excluded_repaired_cells": len(REPAIRED_CELLS),
        "n_compared": len(common),
        "n_ratings": n_ratings,
        "charters_changed": int((deltas != 0).sum()),
        "charters_changed_pct": float((deltas != 0).mean() * 100),
        "ratings_changed": flips,
        "ratings_changed_pct": flips / n_ratings * 100,
        "mean_signed_shift": float(deltas.mean()),
        "max_abs_shift": float(np.abs(deltas).max()),
        "max_condition_mean_shift": float(max(abs(v["mean_shift"]) for v in per_condition.values())),
        "per_condition": per_condition,
    }


# ── Cross-model judge convergence ────────────────────────────────────────────


JUDGE_SHORT = {
    "sonnet": "Claude Sonnet 4.5",
    "sonnet5": "Claude Sonnet 5",
    "gemini": "Gemini 2.5 Pro",
    "gpt4o": "GPT-4o (rerun)",
}

JUDGE_LABELS = {
    "sonnet": "Claude Sonnet 4.5 (Anthropic)",
    "sonnet5": "Claude Sonnet 5 (Anthropic)",
    "gemini": "Gemini 2.5 Pro (Google)",
    "gpt4o": "GPT-4o 2024-11-20 (OpenAI, determinism control)",
}


def judge_convergence(primary_scores: list[dict], matrices: dict, specs: list[str],
                      data_dir: Path) -> dict | None:
    """Compare the primary GPT-4o judge against judges from other vendors.

    Three of the four reviewers (R2 #3, R3 #2, R4) raised the same objection: GPT-4o
    generated the charters and GPT-4o graded them, so the reported ordering could be a
    model preferring its own output rather than a quality difference.

    The test is a rescore, not a regeneration. Every judge sees the identical frozen
    corpus, the identical rubric prompt, and the same temperature; only the judge model
    changes. Two things are then asked of each judge independently:

      1. Does it agree with GPT-4o on individual ratings, and when it disagrees, which
         way does it lean? A judge that is systematically more severe than GPT-4o on
         GPT-4o's own output is direct evidence of the bias the reviewers suspected.
      2. Does the paper's inferential result survive? Absolute scores may move without
         disturbing the ordering, and it is the ordering — the ablation contrasts — that
         the paper's claims rest on. This recomputes each contrast at the specification
         level under the new judge, exactly as the primary analysis does.

    Partial runs are ignored: a judge file that does not carry the full corpus and a
    completion flag contributes nothing, so an interrupted rescore can never leak a
    half-finished number into the manuscript.
    """
    files = sorted(data_dir.glob("etcg-scores-*.json"))
    if not files:
        return None

    def index(rows):
        return {(r["spec_id"], r["condition"], r["charter_id"]): r
                for r in rows if "error" not in r}

    base = index(primary_scores)
    judges: dict[str, dict] = {}
    skipped: list[str] = []

    for path in files:
        slug = path.stem[len("etcg-scores-"):]
        try:
            blob = json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            skipped.append(f"{slug} (unreadable)")
            continue

        meta = blob.get("run_metadata", {})
        other = index(blob.get("scores", []))
        common = sorted(set(base) & set(other))

        if not meta.get("complete"):
            skipped.append(f"{slug} (incomplete: {len(other)}/{len(base)} charters)")
            continue
        if len(common) != len(base):
            skipped.append(f"{slug} (covers {len(common)}/{len(base)} charters)")
            continue

        # Dimension-level ratings, the same unit the human IRR analysis uses.
        pairs = [(base[k]["scores"][d], other[k]["scores"][d])
                 for k in common for d in DIMENSIONS]
        n_ratings = len(pairs)
        exact = sum(1 for a, b in pairs if a == b)
        within_one = sum(1 for a, b in pairs if abs(a - b) <= 1)
        two_point = sum(1 for a, b in pairs if abs(a - b) == 2)
        primary_higher = sum(1 for a, b in pairs if a > b)
        judge_higher = sum(1 for a, b in pairs if b > a)

        two_point_primary_higher = sum(1 for a, b in pairs if a - b == 2)

        per_dimension = {}
        for d in DIMENSIONS:
            dp = [(base[k]["scores"][d], other[k]["scores"][d]) for k in common]
            per_dimension[d] = {
                "exact_pct": sum(1 for a, b in dp if a == b) / len(dp) * 100,
                "primary_mean": float(np.mean([a for a, _ in dp])),
                "judge_mean": float(np.mean([b for _, b in dp])),
            }

        # Charter-level percentages, by condition.
        by_condition = {}
        for cond in CONDITIONS:
            keys = [k for k in common if k[1] == cond]
            j = np.array([other[k]["percentage"] for k in keys], dtype=float)
            b = np.array([base[k]["percentage"] for k in keys], dtype=float)
            by_condition[cond] = {
                "n": len(keys),
                "judge_mean": float(j.mean()),
                "judge_sd": float(j.std(ddof=1)),
                "primary_mean": float(b.mean()),
                "leniency_gap": float(b.mean() - j.mean()),
            }

        # Does the paper's inferential result survive this judge? Recompute the
        # contrasts at the specification level, the way the primary analysis does.
        judge_matrices = build_matrices([other[k] for k in common], specs)
        replication = {}
        for a_cond, b_cond, slug_c, label in CONTRASTS:
            under_judge = paired_contrast(judge_matrices["spec_pct"][a_cond],
                                          judge_matrices["spec_pct"][b_cond], label)
            under_primary = paired_contrast(matrices["spec_pct"][a_cond],
                                            matrices["spec_pct"][b_cond], label)
            same_sign = (under_judge["mean_diff"] >= 0) == (under_primary["mean_diff"] >= 0)
            replication[slug_c] = {
                "label": label,
                "judge": under_judge,
                "primary_mean_diff": under_primary["mean_diff"],
                "primary_p": under_primary["p_signed_rank"],
                "same_sign": bool(same_sign),
                "judge_significant": bool(under_judge["p_signed_rank"] < 0.05),
                "primary_significant": bool(under_primary["p_signed_rank"] < 0.05),
            }

        ranking = sorted(CONDITIONS, key=lambda c: -by_condition[c]["judge_mean"])
        primary_ranking = sorted(CONDITIONS, key=lambda c: -by_condition[c]["primary_mean"])

        overall_j = float(np.mean([other[k]["percentage"] for k in common]))
        overall_p = float(np.mean([base[k]["percentage"] for k in common]))

        judges[slug] = {
            "slug": slug,
            "label": JUDGE_LABELS.get(slug, meta.get("scorer_label", slug)),
            "model": meta.get("scorer_model"),
            "model_resolved": meta.get("scorer_model_resolved"),
            "n_charters": len(common),
            "n_ratings": n_ratings,
            "exact_pct": exact / n_ratings * 100,
            "within_one_pct": within_one / n_ratings * 100,
            "two_point_n": two_point,
            "two_point_pct": two_point / n_ratings * 100,
            "two_point_primary_higher": two_point_primary_higher,
            "primary_higher_n": primary_higher,
            "judge_higher_n": judge_higher,
            "primary_higher_pct": primary_higher / n_ratings * 100,
            "judge_higher_pct": judge_higher / n_ratings * 100,
            "gwet_ac2": gwet_ac2(pairs, [SCALE_MIN, 2, SCALE_MAX]),
            "krippendorff_alpha": krippendorff_alpha_ordinal(pairs, [SCALE_MIN, 2, SCALE_MAX]),
            "overall_judge_mean": overall_j,
            "overall_primary_mean": overall_p,
            "overall_leniency_gap": overall_p - overall_j,
            "per_dimension": per_dimension,
            "by_condition": by_condition,
            "replication": replication,
            "ranking": ranking,
            "ranking_matches_primary": ranking == primary_ranking,
            "cost_and_latency": blob.get("cost_and_latency", {}),
        }

    if not judges:
        return {"judges": {}, "skipped": skipped} if skipped else None

    return {"judges": judges, "skipped": skipped,
            "n_judges": len(judges),
            "all_preserve_ranking": all(j["ranking_matches_primary"] for j in judges.values()),
            "all_more_severe": all(j["overall_leniency_gap"] > 0 for j in judges.values())}


# ── Multi-model generalisation (task 2.3) ───────────────────────────────────


MULTIMODEL_SHORT = {
    "sonnet": "Claude Sonnet 4.5",
    "gemini": "Gemini 2.5 Pro",
    "llama": "Llama 3.3 70B",
}


def multimodel_generalization(mm_dir: Path, primary_matrices: dict,
                              specs: list[str]) -> dict | None:
    """Re-test the ablation pattern with the charter *generator* swapped out.

    Reviewers 1 and 2 (comment 5) note that "model-agnostic" is asserted but only
    GPT-4o is evaluated. multimodel_generate.py re-runs the three conditions on other
    model families; multimodel_score.py scores each set with the same fixed GPT-4o
    rubric judge. This function recomputes the two ablation contrasts (role framing,
    output schema) and the headline contrast for every generator, at the
    specification level, exactly as the primary analysis does, and asks whether the
    direction of each effect is a property of the prompt design or of GPT-4o.

    A generator whose score file is incomplete, unreadable, or missing any
    specification x condition cell contributes nothing.
    """
    if not mm_dir or not mm_dir.exists():
        return None
    slugs = sorted(p.name for p in mm_dir.iterdir() if p.is_dir())
    if not slugs:
        return None

    primary_means = {c: float(np.nanmean(primary_matrices["spec_pct"][c])) for c in CONDITIONS}
    primary_ranking = sorted(CONDITIONS, key=lambda c: -primary_means[c])
    primary_contrasts = {
        slug_c: paired_contrast(primary_matrices["spec_pct"][a],
                                primary_matrices["spec_pct"][b], label)
        for a, b, slug_c, label in CONTRASTS
    }

    models: dict[str, dict] = {}
    skipped: list[str] = []

    for slug in slugs:
        sp = mm_dir / slug / "scores.json"
        if not sp.exists():
            skipped.append(f"{slug} (no scores.json)")
            continue
        try:
            blob = json.loads(sp.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            skipped.append(f"{slug} (unreadable)")
            continue
        meta = blob.get("run_metadata", {})
        if not meta.get("complete"):
            skipped.append(f"{slug} (scoring incomplete)")
            continue

        sc = [s for s in blob.get("scores", []) if "error" not in s]
        cells = Counter((s["spec_id"], s["condition"]) for s in sc)
        empty = [(s, c) for s in specs for c in CONDITIONS if (s, c) not in cells]
        if empty:
            skipped.append(f"{slug} ({len(empty)} empty spec x condition cells)")
            continue

        mats = build_matrices(sc, specs)
        cond_stats = {}
        for c in CONDITIONS:
            v = mats["spec_pct"][c]
            per_cell = [cells[(s, c)] for s in specs]
            cond_stats[c] = {
                "mean": float(v.mean()),
                "sd": float(v.std(ddof=1)),
                "min_charters_per_cell": min(per_cell),
                "mean_charters_per_cell": sum(per_cell) / len(per_cell),
            }

        contrasts = {}
        for a, b, slug_c, label in CONTRASTS:
            res = paired_contrast(mats["spec_pct"][a], mats["spec_pct"][b], label)
            pr = primary_contrasts[slug_c]
            res["same_sign_as_primary"] = bool((res["mean_diff"] >= 0) == (pr["mean_diff"] >= 0))
            res["significant"] = bool(res["p_signed_rank"] < 0.05)
            contrasts[slug_c] = res

        ranking = sorted(CONDITIONS, key=lambda c: -cond_stats[c]["mean"])
        generation = generation_cost(mm_dir / slug / "generation-instrumentation.json")
        generation_cost_usd = generation["total_cost_usd"] if generation else None
        scoring_cost_usd = meta.get("total_cost_usd")
        if not isinstance(scoring_cost_usd, (int, float)):
            scoring_cost_usd = None
        models[slug] = {
            "slug": slug,
            "label": meta.get("generator_label", slug),
            "short": MULTIMODEL_SHORT.get(slug, slug),
            "n_charters": len(sc),
            "condition": cond_stats,
            "contrasts": contrasts,
            "ranking": ranking,
            "ranking_matches_primary": ranking == primary_ranking,
            "baseline_last": ranking[-1] == "baseline",
            "beats_primary_etcg_with_baseline": cond_stats["baseline"]["mean"] > primary_means["etcg"],
            "etcg_minus_primary_etcg_pp": cond_stats["etcg"]["mean"] - primary_means["etcg"],
            "generation_cost_usd": generation_cost_usd,
            "scoring_cost_usd": scoring_cost_usd,
        }

    if not models:
        return {"models": {}, "skipped": skipped, "n_models": 0} if skipped else None

    # Roll-up counts include the primary GPT-4o run as one generator among N.
    def role(md):
        return md["role_guidance"]

    def schema(md):
        return md["schema"]

    prim = {k: {"mean_diff": v["mean_diff"], "p": v["p_signed_rank"],
                "significant": v["p_signed_rank"] < 0.05}
            for k, v in primary_contrasts.items()}

    role_positive = sum(1 for m in models.values() if role(m["contrasts"])["mean_diff"] > 0) \
        + int(prim["role_guidance"]["mean_diff"] > 0)
    role_significant = sum(1 for m in models.values() if role(m["contrasts"])["significant"]) \
        + int(prim["role_guidance"]["significant"])
    schema_nonsig = sum(1 for m in models.values() if not schema(m["contrasts"])["significant"]) \
        + int(not prim["schema"]["significant"])
    schema_negative = sum(1 for m in models.values() if schema(m["contrasts"])["mean_diff"] < 0) \
        + int(prim["schema"]["mean_diff"] < 0)
    baseline_last = sum(1 for m in models.values() if m["baseline_last"]) \
        + int(primary_ranking[-1] == "baseline")
    cost_complete = all(m["generation_cost_usd"] is not None and
                        m["scoring_cost_usd"] is not None for m in models.values())
    total_cost_usd = (sum(m["generation_cost_usd"] + m["scoring_cost_usd"]
                          for m in models.values()) if cost_complete else None)

    n_total = len(models) + 1
    return {
        "models": models,
        "skipped": skipped,
        "n_models": len(models),
        "n_generators": n_total,
        "primary_means": primary_means,
        "primary_ranking": primary_ranking,
        "primary_contrasts": prim,
        "role_positive_in": role_positive,
        "role_significant_in": role_significant,
        "schema_nonsignificant_in": schema_nonsig,
        "schema_negative_in": schema_negative,
        "baseline_last_in": baseline_last,
        "any_beats_primary_etcg_with_baseline": any(
            m["beats_primary_etcg_with_baseline"] for m in models.values()),
        "total_cost_usd": total_cost_usd,
    }


# ── Approach vocabulary ──────────────────────────────────────────────────────


def approach_vocabulary(etcg_results_path: Path) -> dict | None:
    """Characterise the exploratory techniques the model chose for itself.

    Reviewer 1 asked whether the model was given definitions of the exploratory
    techniques it names in the `approach` field. It was not: the schema types that
    field as free text and the prompt supplies no controlled vocabulary. This
    function measures the consequence — how many distinct technique labels the
    model produced, how many it used exactly once, and how much of the spread is
    mere surface variation (case and whitespace) rather than genuinely distinct
    technique choices.
    """
    if not etcg_results_path.exists():
        return None
    with etcg_results_path.open() as fh:
        payload = json.load(fh)

    approaches = [
        charter.get("approach", "")
        for record in payload.get("results", [])
        for charter in record.get("etcg_output", {}).get("charters", [])
    ]
    approaches = [a for a in approaches if a]
    if not approaches:
        return None

    surface = Counter(approaches)
    normalised = Counter(" ".join(a.lower().split()) for a in approaches)

    return {
        "n_charters": len(approaches),
        "n_distinct": len(surface),
        "n_distinct_normalised": len(normalised),
        "n_singletons": sum(1 for _, c in normalised.items() if c == 1),
        "top_share": normalised.most_common(1)[0][1] / len(approaches) * 100.0,
        "top_label": normalised.most_common(1)[0][0],
    }


# The specification traced through all three conditions in the manuscript's
# worked example (Section 4.5). Fixed here so the figures quoted in that section
# are regenerated with everything else rather than transcribed by hand.
WORKED_EXAMPLE_SPEC = "SPEC-23"


def worked_example_verbosity(repo: Path, spec_id: str = WORKED_EXAMPLE_SPEC) -> dict | None:
    """Word counts for one specification's output under each condition.

    The manuscript's worked example contrasts how much text each condition
    produces for the same five charters. These are the numbers it quotes.
    """
    sources = {
        "etcg": ("etcg-results.json", "etcg_output"),
        "intermediate": ("etcg-intermediate-results.json", "intermediate_output"),
        "baseline": ("etcg-baseline-results.json", "baseline_output"),
    }
    out: dict[str, int] = {}
    for cond, (filename, field) in sources.items():
        path = repo / "data" / filename
        if not path.exists():
            return None
        with path.open() as fh:
            records = json.load(fh).get("results", [])
        record = next((r for r in records if r.get("spec_id") == spec_id), None)
        if record is None:
            return None
        payload = record.get(field, {})
        text = payload.get("raw_output")
        if text is None:
            # ETCG output is structured; serialise the charter fields it renders.
            text = " ".join(
                " ".join(str(v) for v in charter.values())
                for charter in payload.get("charters", [])
            )
        out[cond] = len(text.split())
    out["spec_id"] = spec_id
    return out


# ── Generation-side operational cost and latency (task 2.4) ──────────────────


GEN_MODEL_DISPLAY = "GPT-4o"


def generation_cost(path: Path) -> dict | None:
    """Summarise the dedicated generation-instrumentation run (scripts/instrument_generation.py).

    Reviewer 2 comment 6 asks for the framework's operational envelope: latency per
    specification, token consumption, API cost, retry rate, and JSON-validation
    overhead. The primary generation run did not record the API `usage` block, so this
    is measured in a separate run under byte-identical configuration; the evaluation
    corpus is not regenerated.

    An incomplete run, or one that carried any error, returns None — the same posture
    as judge_convergence: a half-finished cost table never reaches the manuscript.
    """
    if not path.exists():
        return None
    try:
        blob = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return None
    if not blob.get("run_metadata", {}).get("complete"):
        return None

    calls = [c for c in blob.get("calls", []) if "error" not in c]
    if not calls:
        return None

    def rows_for(cond: str) -> list[dict]:
        return [c for c in calls if c["condition"] == cond]

    per_condition = {}
    for cond in CONDITIONS:
        rs = rows_for(cond)
        if not rs:
            return None
        lat = sorted(c["latency_s"] for c in rs)
        cin = [c["prompt_tokens"] for c in rs if isinstance(c["prompt_tokens"], (int, float))]
        cout = [c["completion_tokens"] for c in rs if isinstance(c["completion_tokens"], (int, float))]
        cost = [c["cost_usd"] for c in rs if isinstance(c["cost_usd"], (int, float))]
        n = len(rs)
        per_condition[cond] = {
            "n_calls": n,
            "latency_mean": sum(lat) / len(lat),
            "latency_median": lat[len(lat) // 2],
            "latency_max": max(lat),
            "tok_in_total": int(sum(cin)),
            "tok_out_total": int(sum(cout)),
            "tok_in_mean": sum(cin) / len(cin),
            "tok_out_mean": sum(cout) / len(cout),
            "cost_total": sum(cost),
            "cost_per_spec": sum(cost) / n,
            "cost_per_charter": sum(cost) / (n * CHARTERS_PER_CELL),
            "calls_with_retry": sum(1 for c in rs if c.get("retries")),
            "total_retries": sum(c.get("retries", 0) for c in rs),
            "truncated": sum(1 for c in rs if c.get("finish_reason") == "length"),
            "not_five_charters": sum(1 for c in rs if c.get("charter_count") != 5),
            "json_invalid": sum(1 for c in rs if c.get("json_valid") is False),
            "schema_incomplete": sum(1 for c in rs
                                     if c.get("json_valid") and c.get("schema_complete") is False),
            "underextracted": sum(1 for c in rs if c.get("underextracted")),
        }

    all_cost = [c["cost_usd"] for c in calls if isinstance(c["cost_usd"], (int, float))]
    return {
        "per_condition": per_condition,
        "n_calls": len(calls),
        "total_cost_usd": sum(all_cost),
        "model": blob["run_metadata"].get("model"),
        "model_display": GEN_MODEL_DISPLAY,
        "timestamp": blob["run_metadata"].get("timestamp"),
    }


# ── Inter-rater reliability ──────────────────────────────────────────────────


def load_irr(key_path: Path, xlsx_path: Path) -> list[dict] | None:
    """Load the paired automated / human ratings for the IRR sample."""
    if not key_path.exists() or not xlsx_path.exists():
        return None

    with key_path.open(encoding="utf-8") as fh:
        key = json.load(fh)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    human = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if not label.upper().startswith("CH"):
            continue
        vals = [row[i] for i in range(1, 6)]
        if any(v is None for v in vals):
            continue
        human[label.upper()] = [int(v) for v in vals]

    paired = []
    for entry in key:
        n = int(entry["neutral_id"].split("-")[1])
        label = f"CH{n:02d}"
        if label not in human:
            continue
        paired.append({
            "neutral_id": entry["neutral_id"],
            "spec_id": entry["spec_id"],
            "condition": entry["condition"],
            "auto": [int(entry["auto_scores"][d]) for d in DIMENSIONS],
            "human": human[label],
        })
    return paired or None


def krippendorff_alpha_ordinal(pairs: list[tuple[int, int]], categories: list[int]) -> float:
    """Krippendorff's alpha with the ordinal difference metric, two raters, complete data.

    Reported because the round 1 submission leaned on raw within-one-point agreement, which
    Reviewer 2 correctly called permissive. Alpha is chance-corrected but, unlike kappa,
    does not collapse under a skewed marginal distribution in the same way.
    """
    coincidence = defaultdict(float)
    for v1, v2 in pairs:
        coincidence[(v1, v2)] += 1.0
        coincidence[(v2, v1)] += 1.0

    n_c = {c: sum(coincidence[(c, k)] for k in categories) for c in categories}
    n_total = sum(n_c.values())
    if n_total <= 1:
        return float("nan")

    def delta_sq(c: int, k: int) -> float:
        lo, hi = (c, k) if c <= k else (k, c)
        span = sum(n_c[g] for g in categories if lo <= g <= hi)
        return (span - (n_c[c] + n_c[k]) / 2.0) ** 2

    d_obs = sum(coincidence[(c, k)] * delta_sq(c, k) for c in categories for k in categories)
    d_exp = sum(n_c[c] * n_c[k] * delta_sq(c, k) for c in categories for k in categories)
    d_exp /= (n_total - 1)

    if d_exp == 0:
        return float("nan")
    return 1.0 - (d_obs / d_exp)


def gwet_ac2(pairs: list[tuple[int, int]], categories: list[int]) -> float:
    """Gwet's AC2 with linear weights.

    AC2 is used here because it is stable under the high-prevalence ratings that make
    weighted kappa collapse — the kappa paradox the round 1 submission described but
    handled by falling back to raw agreement instead of a chance-corrected statistic.
    """
    q = len(categories)
    if q < 2:
        return float("nan")
    cat_index = {c: i for i, c in enumerate(categories)}
    n = len(pairs)
    if n == 0:
        return float("nan")

    def weight(i: int, k: int) -> float:
        return 1.0 - abs(i - k) / (q - 1)

    p_a = sum(weight(cat_index[v1], cat_index[v2]) for v1, v2 in pairs) / n

    marg = np.zeros(q)
    for v1, v2 in pairs:
        marg[cat_index[v1]] += 0.5
        marg[cat_index[v2]] += 0.5
    marg /= n

    t_w = sum(weight(i, k) for i in range(q) for k in range(q))
    p_e = (t_w / (q * (q - 1))) * sum(p * (1 - p) for p in marg)

    if p_e >= 1.0:
        return float("nan")
    return (p_a - p_e) / (1 - p_e)


def analyse_irr(paired: list[dict]) -> dict:
    """Per-dimension and overall agreement between the automated and human rater."""
    categories = list(range(SCALE_MIN, SCALE_MAX + 1))
    per_dim = {}
    all_pairs: list[tuple[int, int]] = []

    for j, dim in enumerate(DIMENSIONS):
        pairs = [(p["auto"][j], p["human"][j]) for p in paired]
        all_pairs.extend(pairs)
        diffs = np.array([a - h for a, h in pairs], dtype=float)
        absd = np.abs(diffs)
        auto_v = np.array([a for a, _ in pairs], dtype=float)
        human_v = np.array([h for _, h in pairs], dtype=float)

        try:
            kw = float(stats.cohen_kappa_score(auto_v, human_v))  # type: ignore[attr-defined]
        except Exception:
            kw = weighted_kappa(pairs, categories)

        per_dim[dim] = {
            "n": len(pairs),
            "auto_mean": float(auto_v.mean()),
            "human_mean": float(human_v.mean()),
            "delta": float(human_v.mean() - auto_v.mean()),
            "exact_pct": float((absd == 0).mean() * 100),
            "within1_pct": float((absd <= 1).mean() * 100),
            "two_point_n": int((absd >= 2).sum()),
            "alpha_ordinal": krippendorff_alpha_ordinal(pairs, categories),
            "ac2": gwet_ac2(pairs, categories),
            "weighted_kappa": kw,
        }

    absd_all = np.array([abs(a - h) for a, h in all_pairs], dtype=float)
    auto_tot = np.array([sum(p["auto"]) for p in paired], dtype=float)
    human_tot = np.array([sum(p["human"]) for p in paired], dtype=float)

    overall = {
        "n_charters": len(paired),
        "n_ratings": len(all_pairs),
        "exact_pct": float((absd_all == 0).mean() * 100),
        "within1_pct": float((absd_all <= 1).mean() * 100),
        "two_point_n": int((absd_all >= 2).sum()),
        "two_point_pct": float((absd_all >= 2).mean() * 100),
        "auto_mean_pct": float((auto_tot / MAX_CHARTER_POINTS * 100).mean()),
        "human_mean_pct": float((human_tot / MAX_CHARTER_POINTS * 100).mean()),
        "mean_abs_total_diff": float(np.abs(auto_tot - human_tot).mean()),
        "alpha_ordinal": krippendorff_alpha_ordinal(all_pairs, categories),
        "ac2": gwet_ac2(all_pairs, categories),
        "weighted_kappa": weighted_kappa(all_pairs, categories),
        "conditions_covered": sorted({p["condition"] for p in paired}),
        "conditions_missing": sorted(set(CONDITIONS) - {p["condition"] for p in paired}),
    }

    # Locate the two-point disagreements. The round 1 submission asserted none existed.
    locations = []
    for p in paired:
        for j, dim in enumerate(DIMENSIONS):
            if abs(p["auto"][j] - p["human"][j]) >= 2:
                locations.append({
                    "neutral_id": p["neutral_id"], "spec_id": p["spec_id"],
                    "condition": p["condition"], "dimension": dim,
                    "auto": p["auto"][j], "human": p["human"][j],
                })
    overall["two_point_locations"] = locations

    return {"per_dimension": per_dim, "overall": overall}


def weighted_kappa(pairs: list[tuple[int, int]], categories: list[int]) -> float:
    """Linear weighted kappa — reported only to document the kappa paradox."""
    q = len(categories)
    idx = {c: i for i, c in enumerate(categories)}
    n = len(pairs)
    if n == 0:
        return float("nan")
    obs = np.zeros((q, q))
    for v1, v2 in pairs:
        obs[idx[v1], idx[v2]] += 1
    obs /= n
    r_marg = obs.sum(axis=1)
    c_marg = obs.sum(axis=0)
    exp = np.outer(r_marg, c_marg)
    w = np.array([[1 - abs(i - k) / (q - 1) for k in range(q)] for i in range(q)])
    p_o = float((w * obs).sum())
    p_e = float((w * exp).sum())
    if p_e >= 1.0:
        return float("nan")
    return (p_o - p_e) / (1 - p_e)


# ── Round 2: second independent human rater ──────────────────────────────────
#
# Round 1 had one rater score 50 charters covering two of the three conditions.
# Reviewers 2 and 4 asked for more independent evaluation; Reviewer 2 specifically
# noted that the Intermediate condition — where the schema ablation lives — had no
# human validation at all. Round 2 sent a 75-charter packet to a second rater: the
# 25 ETCG + 25 Baseline charters rater 1 already scored (carried over verbatim, so
# rater 1's ratings stay valid and directly comparable) plus 25 new Intermediate
# charters, one per specification. The key file records a `round1_id` on every
# carried-over charter, which is the join for human-vs-human agreement.


def load_irr_r2(key_path: Path, csv_path: Path) -> list[dict] | None:
    """Pair the automated scores with rater 2's ratings for the 75-charter packet.

    Returns records in the same shape ``analyse_irr`` consumes, plus ``round1_id``
    (present only on the ETCG/Baseline charters carried over from round 1).
    """
    if not key_path.exists() or not csv_path.exists():
        return None

    with key_path.open(encoding="utf-8") as fh:
        key = {e["id"]: e for e in json.load(fh)}

    human: dict[str, list[int]] = {}
    with csv_path.open(encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            cid = (row.get("charter_id") or "").strip()
            if not cid.upper().startswith("CHR-R2-"):
                continue
            try:
                vals = [int(row[d]) for d in DIMENSIONS]
            except (KeyError, ValueError, TypeError):
                continue
            if any(v < SCALE_MIN or v > SCALE_MAX for v in vals):
                raise SystemExit(f"{cid}: rater 2 score outside {SCALE_MIN}..{SCALE_MAX}")
            human[cid] = vals

    paired = []
    for cid, entry in key.items():
        if cid not in human:
            continue
        paired.append({
            "neutral_id": cid,
            "round1_id": entry.get("round1_id"),
            "spec_id": entry["spec_id"],
            "condition": entry["condition"],
            "auto": [int(entry["auto_scores"][d]) for d in DIMENSIONS],
            "human": human[cid],
        })
    return paired or None


def analyse_consensus(paired: list[dict]) -> dict:
    """Automated evaluator versus the mean of the two human raters, on the overlap.

    Each record must carry ``auto_scores`` (the automated dimension vector) and the
    two human vectors ``r1`` and ``r2``. The consensus is the per-dimension mean of
    the two humans; the question is whether the automated evaluator sits above it.
    """
    per_dim = {}
    for j, dim in enumerate(DIMENSIONS):
        auto = np.array([p["auto_scores"][j] for p in paired], dtype=float)
        cons = np.array([(p["r1"][j] + p["r2"][j]) / 2.0 for p in paired], dtype=float)
        d = np.abs(auto - cons)
        per_dim[dim] = {
            "auto_mean": float(auto.mean()),
            "consensus_mean": float(cons.mean()),
            "delta": float(auto.mean() - cons.mean()),
            "within_half_pct": float((d <= 0.5).mean() * 100),
            "mean_abs_diff": float(d.mean()),
        }

    auto_tot = np.array([sum(p["auto_scores"]) for p in paired], dtype=float)
    cons_tot = np.array(
        [sum((p["r1"][j] + p["r2"][j]) / 2.0 for j in range(len(DIMENSIONS))) for p in paired],
        dtype=float,
    )
    overall = {
        "n_charters": len(paired),
        "auto_mean_pct": float((auto_tot / MAX_CHARTER_POINTS * 100).mean()),
        "consensus_mean_pct": float((cons_tot / MAX_CHARTER_POINTS * 100).mean()),
        "delta_pct": float(((auto_tot - cons_tot) / MAX_CHARTER_POINTS * 100).mean()),
        "mean_abs_total_diff": float(np.abs(auto_tot - cons_tot).mean()),
    }
    return {"per_dimension": per_dim, "overall": overall}


def build_irr_r2_bundle(paired_r2: list[dict], paired_r1: list[dict] | None) -> dict:
    """Every round 2 agreement analysis, computed from one pair of inputs.

    - ``all``          automated vs rater 2 across all 75 charters (three conditions)
    - ``by_condition`` the same, split by condition — this is where the Intermediate
                       cut lives, the condition round 1 never validated
    - ``human_human``  rater 1 vs rater 2 on the 50-charter overlap (``analyse_irr``
                       shape, with rater 1 in the ``auto`` slot and rater 2 in the
                       ``human`` slot)
    - ``consensus``    automated vs the mean of the two humans on that overlap
    """
    bundle: dict = {
        "all": analyse_irr(paired_r2),
        "by_condition": {
            c: analyse_irr([p for p in paired_r2 if p["condition"] == c])
            for c in CONDITIONS
            if any(p["condition"] == c for p in paired_r2)
        },
    }

    if paired_r1:
        r1_by_nid = {p["neutral_id"]: p for p in paired_r1}
        hh_paired = []
        for p2 in paired_r2:
            nid = p2.get("round1_id")
            if not nid or nid not in r1_by_nid:
                continue
            p1 = r1_by_nid[nid]
            hh_paired.append({
                "neutral_id": nid,
                "spec_id": p1["spec_id"],
                "condition": p1["condition"],
                "auto": p1["human"],          # rater 1 occupies the "auto" slot
                "human": p2["human"],         # rater 2 occupies the "human" slot
                "r1": p1["human"],
                "r2": p2["human"],
                "auto_scores": p1["auto"],    # the automated vector, for consensus
            })
        if hh_paired:
            bundle["human_human"] = analyse_irr(hh_paired)
            bundle["consensus"] = analyse_consensus(hh_paired)
            bundle["n_overlap"] = len(hh_paired)

    return bundle


# ── Registration: compute once, register everything ──────────────────────────


def register_all(reg: Registry, matrices: dict, scores: list[dict],
                 design: dict, irr: dict | None, mixed: dict | None,
                 stability: dict | None = None, vocab: dict | None = None,
                 worked: dict | None = None, judges: dict | None = None,
                 irr_r2: dict | None = None, gencost: dict | None = None,
                 multimodel: dict | None = None, prerepair: dict | None = None) -> None:

    # ---- Corpus descriptives -------------------------------------------------
    reg.put("corpus.n_specs", design["n_specs"], desc="Specifications in the evaluation dataset",
            fmt="int", method="count of distinct spec_id in the scoring file")
    reg.put("corpus.n_conditions", len(CONDITIONS), desc="Experimental conditions", fmt="int",
            method="count of distinct condition values")
    reg.put("corpus.n_charters", design["n_charters"], desc="Total charters scored", fmt="int",
            method="count of scoring records")
    reg.put("corpus.charters_per_condition", design["n_charters"] // len(CONDITIONS),
            desc="Charters per condition", fmt="int", method="total charters / conditions")
    reg.put("corpus.charters_per_spec", CHARTERS_PER_CELL,
            desc="Charters generated per specification per condition", fmt="int",
            method="design constant, verified by validate_design")
    reg.put("corpus.n_sparse", len(SPARSE_SPECS), desc="Sparse (description-only) specifications",
            fmt="int", method="SPEC-01..18")
    reg.put("corpus.n_structured", len(STRUCTURED_SPECS),
            desc="Structured specifications conforming to the ETCG input model", fmt="int",
            method="SPEC-19..25")

    domains = Counter(r["domain"] for r in scores)
    reg.put("corpus.n_domains", len(domains), desc="Application domains represented", fmt="int",
            method="count of distinct domain values")
    reg.put("max_charter_points", MAX_CHARTER_POINTS,
            desc="Maximum achievable score for a single charter", fmt="int",
            method=f"{len(DIMENSIONS)} dimensions x {SCALE_MAX} points")

    # ---- Approach vocabulary (Reviewer 1: were techniques defined for the model?) --
    if vocab:
        m = ("distinct values of the free-text `approach` field across the "
             f"{vocab['n_charters']} ETCG charters in data/etcg-results.json")
        reg.put("vocab.n_distinct", vocab["n_distinct"],
                desc="Distinct exploratory technique labels produced by the model",
                fmt="int", method=m + ", exact string match")
        reg.put("vocab.n_distinct_normalised", vocab["n_distinct_normalised"],
                desc="Distinct technique labels after case and whitespace normalisation",
                fmt="int", method=m + ", lowercased with whitespace collapsed")
        reg.put("vocab.n_singletons", vocab["n_singletons"],
                desc="Technique labels used exactly once in the corpus",
                fmt="int", method=m + ", normalised, count of labels with frequency 1")
        reg.put("vocab.top_share", vocab["top_share"],
                desc="Share of ETCG charters using the single most common technique (%)",
                method=m + f", normalised; most common label is {vocab['top_label']!r}")

    # ---- Worked example verbosity (Section 4.5) ------------------------------
    if worked:
        m = (f"whitespace-delimited word count of the raw model output for "
             f"{worked['spec_id']} in data/etcg-*-results.json")
        for cond in CONDITIONS:
            reg.put(f"worked.{cond}_words", worked[cond],
                    desc=f"{CONDITION_LABELS[cond]} output length for the worked "
                         f"example specification (words)",
                    fmt="int", method=m)

    # ---- RQ1: overall quality by condition -----------------------------------
    for cond in CONDITIONS:
        spec_vals = matrices["spec_pct"][cond]
        charter_vals = matrices["charter_pct"][cond]
        lo, hi = ci_mean(spec_vals)
        label = CONDITION_LABELS[cond]

        reg.put(f"rq1.{cond}.mean", float(spec_vals.mean()),
                desc=f"{label} mean charter quality (%)",
                method="mean of per-specification means (n=25)")
        reg.put(f"rq1.{cond}.sd_spec", float(spec_vals.std(ddof=1)),
                desc=f"{label} SD of per-specification means (%)",
                method="sample SD across the 25 specification means")
        reg.put(f"rq1.{cond}.ci_low", lo, desc=f"{label} 95% CI lower bound (%)",
                method="t-based CI on the specification-level mean, df=24")
        reg.put(f"rq1.{cond}.ci_high", hi, desc=f"{label} 95% CI upper bound (%)",
                method="t-based CI on the specification-level mean, df=24")
        reg.put(f"rq1.{cond}.sd_charter", float(charter_vals.std(ddof=1)),
                desc=f"{label} SD across individual charters (%) — descriptive only",
                method="sample SD across 125 charter scores; descriptive, not an inferential input")
        reg.put(f"rq1.{cond}.min", float(charter_vals.min()),
                desc=f"{label} lowest single charter score (%)", fmt="{:.1f}",
                method="minimum of 125 charter percentages")
        ceiling_n = int((charter_vals >= 100.0).sum())
        reg.put(f"rq1.{cond}.ceiling_n", ceiling_n,
                desc=f"{label} charters at the maximum rubric score", fmt="int",
                method=f"count of charters scoring {MAX_CHARTER_POINTS}/{MAX_CHARTER_POINTS}")
        reg.put(f"rq1.{cond}.ceiling_pct", ceiling_n / len(charter_vals) * 100,
                desc=f"{label} share of charters at ceiling (%)", fmt="pct0",
                method="ceiling count / 125")

    ceilings = [reg.get(f"rq1.{c}.ceiling_pct") for c in CONDITIONS]
    reg.put("rq1.ceiling_min", min(ceilings), desc="Lowest ceiling share across conditions (%)",
            fmt="pct0", method="minimum of the three ceiling percentages")
    reg.put("rq1.ceiling_max", max(ceilings), desc="Highest ceiling share across conditions (%)",
            fmt="pct0", method="maximum of the three ceiling percentages")

    # ---- RQ1: pairwise contrasts ---------------------------------------------
    for a, b, slug, label in CONTRASTS:
        res = paired_contrast(matrices["spec_pct"][a], matrices["spec_pct"][b], label)
        old = superseded_unpaired(matrices["charter_pct"][a], matrices["charter_pct"][b])
        base = f"rq1.{slug}"
        m = ("Wilcoxon signed-rank on 25 paired per-specification means; "
             "the paired unit is the specification")

        reg.put(f"{base}.mean_diff", res["mean_diff"],
                desc=f"{label}: mean difference (pp)", fmt="diff", method=m)
        reg.put(f"{base}.ci_low", res["ci_low"], desc=f"{label}: 95% CI lower bound (pp)",
                fmt="diff", method="t-based CI on the 25 paired differences")
        reg.put(f"{base}.ci_high", res["ci_high"], desc=f"{label}: 95% CI upper bound (pp)",
                fmt="diff", method="t-based CI on the 25 paired differences")
        reg.put(f"{base}.w", res["w_stat"], desc=f"{label}: Wilcoxon signed-rank W",
                fmt="{:.1f}", method=m)
        reg.put(f"{base}.p", res["p_signed_rank"], desc=f"{label}: p-value (primary test)",
                fmt="p", method=m)
        reg.put(f"{base}.p_t", res["p_paired_t"], desc=f"{label}: paired t-test p (robustness)",
                fmt="p", method="paired t-test on the same 25 differences")
        reg.put(f"{base}.dz", res["dz"], desc=f"{label}: Cohen's dz",
                fmt="diff3", method="mean of paired differences / SD of paired differences")
        reg.put(f"{base}.rb", res["rank_biserial"],
                desc=f"{label}: matched-pairs rank-biserial correlation", fmt="{:.3f}",
                method="(W+ - W-)/(W+ + W-) over non-zero differences")
        reg.put(f"{base}.n_zero", res["n_zero_diff"],
                desc=f"{label}: specifications with zero difference", fmt="int",
                method="count of tied pairs")
        reg.put(f"{base}.p_superseded", old["p_value"],
                desc=f"{label}: superseded charter-level Mann-Whitney p "
                     f"(round 1 method, recomputed on the current repaired data)",
                fmt="p",
                method="Mann-Whitney U on 125 vs 125 charters — pseudoreplicated, reported "
                       "only for comparison in the response letter; the round 1 submission's "
                       "test re-run on the corrected scoring file, not the value it printed")

    # ---- RQ1: pre-repair headline (superseded; Threats provenance paragraph) ------
    # What the baseline-splitter repair changed, recomputed from the retained pre-repair
    # scoring file with the same estimator as above so the disclosure is checkable.
    if prerepair:
        pm = (f"same Wilcoxon signed-rank on 25 paired specification means as the primary "
              f"analysis, recomputed on data/{PREREPAIR_SCORES_NAME} (the scoring file "
              f"before scripts/repair_baseline_split.py)")
        reg.put("rq1.baseline.mean_prerepair", prerepair["conditions"]["baseline"]["mean"],
                desc="Baseline mean charter quality before the splitter repair (%)", method=pm)
        reg.put("rq1.baseline.sd_charter_prerepair",
                prerepair["conditions"]["baseline"]["sd_charter"],
                desc="Baseline SD across individual charters before the splitter repair (%)",
                method=pm)
        for _a, _b, slug, label in CONTRASTS:
            c = prerepair["contrasts"][slug]
            reg.put(f"rq1.{slug}.mean_diff_prerepair", c["mean_diff"],
                    desc=f"{label}: mean difference before the splitter repair (pp)",
                    fmt="diff", method=pm)
            reg.put(f"rq1.{slug}.p_prerepair", c["p_signed_rank"],
                    desc=f"{label}: signed-rank p before the splitter repair", fmt="p",
                    method=pm)
            reg.put(f"rq1.{slug}.ci_low_prerepair", c["ci_low"],
                    desc=f"{label}: 95% CI lower bound before the splitter repair (pp)",
                    fmt="diff", method=pm)
            reg.put(f"rq1.{slug}.ci_high_prerepair", c["ci_high"],
                    desc=f"{label}: 95% CI upper bound before the splitter repair (pp)",
                    fmt="diff", method=pm)

    # ---- RQ2: per-dimension --------------------------------------------------
    dim_pvals = {}
    for dim in DIMENSIONS:
        for cond in CONDITIONS:
            vals = matrices["spec_dim"][cond][dim]
            reg.put(f"rq2.{dim}.{cond}.mean", float(vals.mean()),
                    desc=f"{DIMENSION_LABELS[dim]} mean for {CONDITION_LABELS[cond]} (1-3 scale)",
                    fmt="{:.3f}", method="mean of per-specification dimension means")
        for a, b, slug, label in CONTRASTS:
            res = paired_contrast(matrices["spec_dim"][a][dim], matrices["spec_dim"][b][dim], label)
            reg.put(f"rq2.{dim}.{slug}.p", res["p_signed_rank"],
                    desc=f"{DIMENSION_LABELS[dim]}, {label}: p-value", fmt="p",
                    method="Wilcoxon signed-rank on 25 paired per-specification dimension means")
            reg.put(f"rq2.{dim}.{slug}.diff", res["mean_diff"],
                    desc=f"{DIMENSION_LABELS[dim]}, {label}: mean difference", fmt="diff3",
                    method="difference of per-specification dimension means")
            dim_pvals[f"{dim}.{slug}"] = res["p_signed_rank"]

    adjusted = holm_bonferroni(dim_pvals)
    for k, v in adjusted.items():
        dim, slug = k.split(".")
        reg.put(f"rq2.{dim}.{slug}.p_holm", v,
                desc=f"{DIMENSION_LABELS[dim]}, {slug}: Holm-adjusted p", fmt="p",
                method=f"Holm-Bonferroni across the {len(dim_pvals)} dimension-level tests")
    reg.put("rq2.n_tests", len(dim_pvals), desc="Dimension-level tests in the family", fmt="int",
            method="5 dimensions x 3 contrasts")
    reg.put("rq2.n_sig_raw", sum(1 for p in dim_pvals.values() if p < 0.05),
            desc="Dimension-level tests significant before correction", fmt="int",
            method="count of unadjusted p < 0.05")
    reg.put("rq2.n_sig_holm", sum(1 for p in adjusted.values() if p < 0.05),
            desc="Dimension-level tests significant after Holm correction", fmt="int",
            method="count of Holm-adjusted p < 0.05")

    # ---- RQ3: inter-rater reliability ----------------------------------------
    if irr:
        ov = irr["overall"]
        reg.put("rq3.n_charters", ov["n_charters"], desc="Charters in the IRR sample", fmt="int",
                method="paired records in the IRR key with a completed human score row")
        reg.put("rq3.n_ratings", ov["n_ratings"], desc="Individual dimension ratings compared",
                fmt="int", method="IRR charters x 5 dimensions")
        reg.put("rq3.sample_pct", ov["n_charters"] / design["n_charters"] * 100,
                desc="IRR sample as a share of the full charter corpus (%)", fmt="{:.1f}",
                method="IRR charters / total charters scored")
        reg.put("rq3.exact_pct", ov["exact_pct"], desc="Exact agreement (%)", fmt="{:.1f}",
                method="share of ratings where both raters gave the same score")
        reg.put("rq3.within1_pct", ov["within1_pct"], desc="Within-one-point agreement (%)",
                fmt="{:.1f}", method="share of ratings differing by at most one scale point")
        reg.put("rq3.two_point_n", ov["two_point_n"],
                desc="Ratings differing by two scale points", fmt="int",
                method="count of |auto - human| >= 2 — the round 1 submission asserted zero")
        reg.put("rq3.two_point_pct", ov["two_point_pct"],
                desc="Share of ratings differing by two scale points (%)", fmt="{:.1f}",
                method="two-point count / total ratings")
        reg.put("rq3.alpha", ov["alpha_ordinal"],
                desc="Krippendorff's alpha (ordinal metric)", fmt="{:.3f}",
                method="ordinal difference metric, two raters, complete data")
        reg.put("rq3.ac2", ov["ac2"], desc="Gwet's AC2 (linear weights)", fmt="{:.3f}",
                method="chance-corrected agreement robust to skewed marginals")
        reg.put("rq3.kappa_w", ov["weighted_kappa"],
                desc="Linear weighted kappa (reported to document the kappa paradox)",
                fmt="{:.3f}", method="linear weights; collapses under high prevalence")
        reg.put("rq3.auto_mean_pct", ov["auto_mean_pct"],
                desc="Automated evaluator mean on the IRR sample (%)",
                method="mean charter percentage over IRR sample, automated scores")
        reg.put("rq3.human_mean_pct", ov["human_mean_pct"],
                desc="Human reviewer mean on the IRR sample (%)",
                method="mean charter percentage over IRR sample, human scores")
        reg.put("rq3.mean_delta_pct", ov["human_mean_pct"] - ov["auto_mean_pct"],
                desc="Human minus automated mean on the IRR sample (pp)", fmt="diff",
                method="difference of the two IRR sample means")
        reg.put("rq3.mean_abs_total_diff", ov["mean_abs_total_diff"],
                desc="Mean absolute difference in total charter score (of 15)",
                method="mean |auto total - human total| across IRR charters")
        reg.put("rq3.n_conditions_covered", len(ov["conditions_covered"]),
                desc="Conditions represented in the IRR sample", fmt="int",
                method="distinct conditions present — a known round 1 gap")

        for dim in DIMENSIONS:
            pd_ = irr["per_dimension"][dim]
            for field, fmt, desc in [
                ("auto_mean", "{:.2f}", "automated mean"),
                ("human_mean", "{:.2f}", "human mean"),
                ("delta", "diff", "human minus automated"),
                ("exact_pct", "{:.1f}", "exact agreement (%)"),
                ("within1_pct", "{:.1f}", "within-one agreement (%)"),
                ("alpha_ordinal", "{:.3f}", "Krippendorff's alpha"),
                ("ac2", "{:.3f}", "Gwet's AC2"),
            ]:
                reg.put(f"rq3.{dim}.{field}", pd_[field],
                        desc=f"{DIMENSION_LABELS[dim]}: {desc}", fmt=fmt,
                        method="automated versus human ratings on the IRR sample")
            reg.put(f"rq3.{dim}.two_point_n", pd_["two_point_n"],
                    desc=f"{DIMENSION_LABELS[dim]}: two-point disagreements", fmt="int",
                    method="count of |auto - human| >= 2 on this dimension")

    # ---- RQ3 (round 2): second independent human rater ---------------------
    if irr_r2:
        _R2_DIM_FIELDS = [
            ("auto_mean", "{:.2f}", "automated mean"),
            ("human_mean", "{:.2f}", "rater 2 mean"),
            ("delta", "diff", "rater 2 minus automated"),
            ("exact_pct", "{:.1f}", "exact agreement (%)"),
            ("within1_pct", "{:.1f}", "within-one agreement (%)"),
            ("alpha_ordinal", "{:.3f}", "Krippendorff's alpha"),
            ("ac2", "{:.3f}", "Gwet's AC2"),
        ]

        def _reg_auto_human(prefix: str, res: dict, scope: str, per_dim: bool) -> None:
            ov = res["overall"]
            for slug, val, fmt, desc in [
                ("n_charters", ov["n_charters"], "int", "charters compared"),
                ("n_ratings", ov["n_ratings"], "int", "dimension ratings compared"),
                ("exact_pct", ov["exact_pct"], "{:.1f}", "exact agreement (%)"),
                ("within1_pct", ov["within1_pct"], "{:.1f}", "within-one-point agreement (%)"),
                ("two_point_n", ov["two_point_n"], "int", "two-point disagreements"),
                ("two_point_pct", ov["two_point_pct"], "{:.1f}",
                 "two-point disagreement rate (%)"),
                ("alpha", ov["alpha_ordinal"], "{:.3f}", "Krippendorff's alpha (ordinal)"),
                ("ac2", ov["ac2"], "{:.3f}", "Gwet's AC2 (linear weights)"),
                ("kappa_w", ov["weighted_kappa"], "{:.3f}", "linear weighted kappa"),
                ("auto_mean_pct", ov["auto_mean_pct"], "{:.1f}",
                 "automated evaluator mean (%)"),
                ("human_mean_pct", ov["human_mean_pct"], "{:.1f}", "rater 2 mean (%)"),
                ("mean_delta_pct", ov["human_mean_pct"] - ov["auto_mean_pct"], "diff",
                 "rater 2 minus automated mean (pp)"),
                ("mean_abs_total_diff", ov["mean_abs_total_diff"], "{:.2f}",
                 "mean |automated total - rater 2 total| (of 15)"),
                ("n_conditions_covered", len(ov["conditions_covered"]), "int",
                 "conditions represented"),
            ]:
                reg.put(f"{prefix}.{slug}", val, fmt=fmt,
                        desc=f"Round 2 IRR, {scope}: {desc}",
                        method="automated evaluator versus rater 2 on the round 2 packet")
            if not per_dim:
                return
            for dim in DIMENSIONS:
                pd_ = res["per_dimension"][dim]
                for field, fmt, d in _R2_DIM_FIELDS:
                    reg.put(f"{prefix}.{dim}.{field}", pd_[field], fmt=fmt,
                            desc=f"Round 2 IRR, {scope}, {DIMENSION_LABELS[dim]}: {d}",
                            method="automated evaluator versus rater 2")
                reg.put(f"{prefix}.{dim}.two_point_n", pd_["two_point_n"], fmt="int",
                        desc=(f"Round 2 IRR, {scope}, {DIMENSION_LABELS[dim]}: "
                              f"two-point disagreements"),
                        method="count of |automated - rater 2| >= 2 on this dimension")

        _reg_auto_human("rq3.r2", irr_r2["all"], "all conditions (n=75)", per_dim=True)
        for cond, res in irr_r2["by_condition"].items():
            _reg_auto_human(f"rq3.r2.{cond}", res, f"{CONDITION_LABELS[cond]} only",
                            per_dim=False)

        if "human_human" in irr_r2:
            hh = irr_r2["human_human"]
            ov = hh["overall"]
            for slug, val, fmt, desc in [
                ("n_charters", ov["n_charters"], "int",
                 "charters on the round 1 / round 2 overlap"),
                ("n_ratings", ov["n_ratings"], "int", "dimension ratings compared"),
                ("exact_pct", ov["exact_pct"], "{:.1f}",
                 "exact agreement between the two human raters (%)"),
                ("within1_pct", ov["within1_pct"], "{:.1f}",
                 "within-one-point agreement between the two human raters (%)"),
                ("two_point_n", ov["two_point_n"], "int",
                 "two-point disagreements between the two human raters"),
                ("two_point_pct", ov["two_point_pct"], "{:.1f}",
                 "two-point disagreement rate between the two human raters (%)"),
                ("alpha", ov["alpha_ordinal"], "{:.3f}",
                 "Krippendorff's alpha (ordinal), rater 1 vs rater 2"),
                ("ac2", ov["ac2"], "{:.3f}", "Gwet's AC2 (linear weights), rater 1 vs rater 2"),
                ("kappa_w", ov["weighted_kappa"], "{:.3f}",
                 "linear weighted kappa, rater 1 vs rater 2"),
                ("r1_mean_pct", ov["auto_mean_pct"], "{:.1f}", "rater 1 mean on the overlap (%)"),
                ("r2_mean_pct", ov["human_mean_pct"], "{:.1f}", "rater 2 mean on the overlap (%)"),
                ("mean_delta_pct", ov["human_mean_pct"] - ov["auto_mean_pct"], "diff",
                 "rater 2 minus rater 1 mean (pp)"),
                ("mean_abs_total_diff", ov["mean_abs_total_diff"], "{:.2f}",
                 "mean |rater 1 total - rater 2 total| (of 15)"),
            ]:
                reg.put(f"rq3.hh.{slug}", val, fmt=fmt,
                        desc=f"Human-human agreement: {desc}",
                        method=("rater 1 versus rater 2 on the 50-charter round 1 / "
                                "round 2 overlap"))
            for dim in DIMENSIONS:
                pd_ = hh["per_dimension"][dim]
                for field, fmt, d in [
                    ("auto_mean", "{:.2f}", "rater 1 mean"),
                    ("human_mean", "{:.2f}", "rater 2 mean"),
                    ("delta", "diff", "rater 2 minus rater 1"),
                    ("exact_pct", "{:.1f}", "exact agreement (%)"),
                    ("within1_pct", "{:.1f}", "within-one agreement (%)"),
                    ("alpha_ordinal", "{:.3f}", "Krippendorff's alpha"),
                    ("ac2", "{:.3f}", "Gwet's AC2"),
                ]:
                    reg.put(f"rq3.hh.{dim}.{field}", pd_[field], fmt=fmt,
                            desc=f"Human-human, {DIMENSION_LABELS[dim]}: {d}",
                            method="rater 1 versus rater 2 on the overlap")
                reg.put(f"rq3.hh.{dim}.two_point_n", pd_["two_point_n"], fmt="int",
                        desc=(f"Human-human, {DIMENSION_LABELS[dim]}: two-point "
                              f"disagreements"),
                        method="count of |rater 1 - rater 2| >= 2 on this dimension")

        if "consensus" in irr_r2:
            cons = irr_r2["consensus"]["overall"]
            for slug, val, fmt, desc in [
                ("n_charters", cons["n_charters"], "int", "charters (two-human overlap)"),
                ("auto_mean_pct", cons["auto_mean_pct"], "{:.1f}",
                 "automated evaluator mean (%)"),
                ("consensus_mean_pct", cons["consensus_mean_pct"], "{:.1f}",
                 "two-rater consensus mean (%)"),
                ("mean_delta_pct", cons["delta_pct"], "diff",
                 "automated minus consensus mean (pp)"),
                ("mean_abs_total_diff", cons["mean_abs_total_diff"], "{:.2f}",
                 "mean |automated total - consensus total| (of 15)"),
            ]:
                reg.put(f"rq3.cons.{slug}", val, fmt=fmt,
                        desc=f"Automated vs two-rater consensus: {desc}",
                        method=("automated evaluator versus the per-dimension mean of "
                                "rater 1 and rater 2"))

    # ---- RQ4: specification richness -----------------------------------------
    specs = matrices["specs"]
    sparse_idx = [i for i, s in enumerate(specs) if s in SPARSE_SPECS]
    struct_idx = [i for i, s in enumerate(specs) if s in STRUCTURED_SPECS]

    rq4_pvals: dict[str, float] = {}
    for cond_i, cond in enumerate(CONDITIONS):
        spec_vals = matrices["spec_pct"][cond]
        for group_i, (group, idxs) in enumerate([("sparse", sparse_idx), ("structured", struct_idx)]):
            vals = spec_vals[idxs]
            reg.put(f"rq4.{cond}.{group}.mean", float(vals.mean()),
                    desc=f"{CONDITION_LABELS[cond]} mean on {group} specifications (%)",
                    method=f"mean of per-specification means, {group} group (n={len(idxs)})")
            reg.put(f"rq4.{cond}.{group}.sd", float(vals.std(ddof=1)),
                    desc=f"{CONDITION_LABELS[cond]} SD on {group} specifications (%)",
                    method=f"SD across the {len(idxs)} specification means in the {group} group")
            reg.put(f"rq4.{cond}.{group}.n", len(idxs),
                    desc=f"Specifications in the {group} group", fmt="int",
                    method="group membership by spec ID")
            ci_low, ci_high = bootstrap_ci_mean(vals, seed=10 * cond_i + group_i)
            reg.put(f"rq4.{cond}.{group}.ci_low", ci_low,
                    desc=f"{CONDITION_LABELS[cond]} bootstrap 95% CI lower bound on {group} specifications (%)",
                    method="percentile bootstrap of specification means (10,000 deterministic resamples)")
            reg.put(f"rq4.{cond}.{group}.ci_high", ci_high,
                    desc=f"{CONDITION_LABELS[cond]} bootstrap 95% CI upper bound on {group} specifications (%)",
                    method="percentile bootstrap of specification means (10,000 deterministic resamples)")

        a = spec_vals[struct_idx]
        b = spec_vals[sparse_idx]
        u, p = stats.mannwhitneyu(a, b, alternative="two-sided")
        pooled = math.sqrt(((len(a) - 1) * a.var(ddof=1) + (len(b) - 1) * b.var(ddof=1))
                           / (len(a) + len(b) - 2))
        reg.put(f"rq4.{cond}.contrast_diff", float(a.mean() - b.mean()),
                desc=f"{CONDITION_LABELS[cond]}: structured minus sparse (pp)", fmt="diff",
                method="difference of group means at the specification level")
        reg.put(f"rq4.{cond}.contrast_p", float(p),
                desc=f"{CONDITION_LABELS[cond]}: structured versus sparse raw p", fmt="p",
                method="Mann-Whitney U — these are independent groups of specifications, "
                       "so an unpaired test is correct here")
        rq4_pvals[cond] = float(p)
        reg.put(f"rq4.{cond}.contrast_d", float((a.mean() - b.mean()) / pooled) if pooled else float("nan"),
                desc=f"{CONDITION_LABELS[cond]}: structured versus sparse Cohen's d",
                fmt="{:.2f}", method="difference of means / pooled SD, specification level")

    rq4_adjusted = holm_bonferroni(rq4_pvals)
    reg.put("rq4.n_tests", len(rq4_pvals), desc="RQ4 condition-wise richness contrasts", fmt="int",
            method="three Mann–Whitney tests form one RQ4 family")
    reg.put("rq4.n_sig_holm", sum(p < 0.05 for p in rq4_adjusted.values()),
            desc="RQ4 contrasts significant after Holm correction", fmt="int",
            method="count of Holm-adjusted p < 0.05")
    for cond, p_holm in rq4_adjusted.items():
        reg.put(f"rq4.{cond}.contrast_p_holm", p_holm,
                desc=f"{CONDITION_LABELS[cond]}: structured versus sparse Holm-adjusted p", fmt="p",
                method=f"Holm-Bonferroni across {len(rq4_adjusted)} RQ4 condition-wise contrasts")

    for dim in DIMENSIONS:
        for group, idxs in [("sparse", sparse_idx), ("structured", struct_idx)]:
            vals = matrices["spec_dim"]["etcg"][dim][idxs]
            reg.put(f"rq4.dim.{dim}.{group}", float(vals.mean()),
                    desc=f"ETCG {DIMENSION_LABELS[dim]} mean on {group} specifications",
                    fmt="{:.3f}",
                    method=f"mean of per-specification dimension means, {group} group")
        reg.put(f"rq4.dim.{dim}.delta",
                float(matrices["spec_dim"]["etcg"][dim][struct_idx].mean()
                      - matrices["spec_dim"]["etcg"][dim][sparse_idx].mean()),
                desc=f"ETCG {DIMENSION_LABELS[dim]}: structured minus sparse", fmt="diff3",
                method="difference of group means at the specification level")

    # ---- Assumption checks ----------------------------------------------------
    for cond in CONDITIONS:
        w, p = stats.shapiro(matrices["spec_pct"][cond])
        reg.put(f"checks.shapiro.{cond}.w", float(w),
                desc=f"Shapiro-Wilk W for {CONDITION_LABELS[cond]} specification means",
                fmt="{:.3f}", method="normality check on the 25 specification-level means")
        reg.put(f"checks.shapiro.{cond}.p", float(p),
                desc=f"Shapiro-Wilk p for {CONDITION_LABELS[cond]} specification means", fmt="p",
                method="normality check on the 25 specification-level means")

    for a, b, slug, label in CONTRASTS:
        diff = matrices["spec_pct"][a] - matrices["spec_pct"][b]
        w, p = stats.shapiro(diff)
        reg.put(f"checks.shapiro_diff.{slug}.p", float(p),
                desc=f"Shapiro-Wilk p on the paired differences, {label}", fmt="p",
                method="normality of paired differences — the assumption the paired t-test needs")

    # ---- Automated scorer stability across runs -------------------------------
    if stability:
        m = ("comparison of two GPT-4o scoring runs at temperature 0 over the same charters "
             "(24 and 25 March 2026)")
        reg.put("stability.n_compared", stability["n_compared"],
                desc="Charters scored in both runs", fmt="int", method=m)
        reg.put("stability.n_ratings", stability["n_ratings"],
                desc="Dimension ratings compared across runs", fmt="int", method=m)
        reg.put("stability.charters_changed", stability["charters_changed"],
                desc="Charters whose total score changed between runs", fmt="int", method=m)
        reg.put("stability.charters_changed_pct", stability["charters_changed_pct"],
                desc="Share of charters whose score changed between runs (%)", fmt="{:.1f}",
                method=m)
        reg.put("stability.ratings_changed", stability["ratings_changed"],
                desc="Individual dimension ratings that changed between runs", fmt="int",
                method=m)
        reg.put("stability.ratings_changed_pct", stability["ratings_changed_pct"],
                desc="Share of dimension ratings that changed between runs (%)", fmt="{:.1f}",
                method=m)
        reg.put("stability.max_abs_shift", stability["max_abs_shift"],
                desc="Largest single-charter score change between runs (pp)", fmt="{:.1f}",
                method=m)
        reg.put("stability.max_condition_mean_shift", stability["max_condition_mean_shift"],
                desc="Largest condition-level mean shift between runs (pp)", fmt="{:.2f}",
                method=m + "; this is the aggregate stability the manuscript claimed")

    # ---- Mixed-effects model --------------------------------------------------
    if mixed and "error" not in mixed:
        m = ("linear mixed-effects model on all 375 charter scores, "
             "pct ~ condition with a random intercept per specification, REML")
        reg.put("mixed.intercept", mixed["intercept"],
                desc="Mixed model intercept — Baseline mean (%)", method=m)
        for _, _, slug, label in CONTRASTS:
            vals = mixed["contrasts"][slug]
            reg.put(f"mixed.{slug}.coef", vals["coef"],
                    desc=f"Mixed model estimate, {label} (pp)", fmt="diff", method=m)
            reg.put(f"mixed.{slug}.se", vals["se"],
                    desc=f"Mixed model standard error, {label}", fmt="{:.3f}", method=m)
            reg.put(f"mixed.{slug}.z", vals["z"],
                    desc=f"Mixed model z, {label}", fmt="{:.3f}", method=m)
            reg.put(f"mixed.{slug}.p", vals["p"],
                    desc=f"Mixed model p, {label}", fmt="p", method=m + ", Wald test")
        reg.put("mixed.group_var", mixed["group_var"],
                desc="Between-specification variance component",
                fmt="{:.3f}", method="random intercept variance from the model")
        reg.put("mixed.resid_var", mixed["resid_var"], desc="Residual (within-specification) variance",
                fmt="{:.3f}", method="model residual scale")
        reg.put("mixed.icc", mixed["icc"],
                desc="Intraclass correlation — the clustering Reviewer 2 identified",
                fmt="{:.3f}",
                method="group variance / (group variance + residual variance); quantifies how "
                       "much the round 1 charter-level analysis overstated its effective n")
        reg.put("mixed.n_obs", mixed["n_obs"], desc="Observations in the mixed model", fmt="int",
                method="all individual charter scores")
        reg.put("mixed.n_groups", mixed["n_groups"], desc="Grouping units (specifications)",
                fmt="int", method="distinct spec_id values")

    # ---- Cross-model judge convergence ----------------------------------------
    if judges and judges.get("judges"):
        for slug, j in sorted(judges["judges"].items()):
            src = (f"rescore of the identical frozen corpus by {j['label']} "
                   f"({j['model']}) at temperature 0 with the primary rubric prompt; "
                   f"only the judge model differs from the primary run")
            k = f"judge.{slug}"

            reg.put(f"{k}.n_charters", j["n_charters"], fmt="int",
                    desc=f"Charters rescored by {j['label']}", method=src)
            reg.put(f"{k}.n_ratings", j["n_ratings"], fmt="int",
                    desc=f"Dimension ratings compared, GPT-4o vs {j['label']}",
                    method=src + "; 5 dimensions per charter")
            reg.put(f"{k}.exact", j["exact_pct"],
                    desc=f"Exact agreement with GPT-4o, {j['label']} (%)", method=src)
            reg.put(f"{k}.within_one", j["within_one_pct"],
                    desc=f"Within-one agreement with GPT-4o, {j['label']} (%)", method=src)
            reg.put(f"{k}.two_point_n", j["two_point_n"], fmt="int",
                    desc=f"Two-point disagreements with GPT-4o, {j['label']}", method=src)
            reg.put(f"{k}.two_point_pct", j["two_point_pct"],
                    desc=f"Two-point disagreement rate, {j['label']} (%)", method=src)
            reg.put(f"{k}.two_point_primary_higher", j["two_point_primary_higher"], fmt="int",
                    desc=("Two-point disagreements where GPT-4o is the higher rater, "
                          f"{j['label']}"),
                    method=src + "; the direction test for same-model bias")
            reg.put(f"{k}.primary_higher_n", j["primary_higher_n"], fmt="int",
                    desc=f"Ratings where GPT-4o scores above {j['label']}", method=src)
            reg.put(f"{k}.judge_higher_n", j["judge_higher_n"], fmt="int",
                    desc=f"Ratings where {j['label']} scores above GPT-4o", method=src)
            reg.put(f"{k}.primary_higher_pct", j["primary_higher_pct"],
                    desc=f"Share of ratings where GPT-4o is more lenient than {j['label']} (%)",
                    method=src)
            reg.put(f"{k}.ac2", j["gwet_ac2"], fmt="{:.3f}",
                    desc=f"Gwet's AC2, GPT-4o vs {j['label']}",
                    method=src + "; ordinal weighting, same estimator as the human IRR")
            reg.put(f"{k}.alpha", j["krippendorff_alpha"], fmt="{:.3f}",
                    desc=f"Krippendorff's alpha, GPT-4o vs {j['label']}",
                    method=src + "; ordinal difference function")
            reg.put(f"{k}.overall_mean", j["overall_judge_mean"],
                    desc=f"Overall charter mean under {j['label']} (%)", method=src)
            reg.put(f"{k}.leniency_gap", j["overall_leniency_gap"], fmt="diff",
                    desc=(f"GPT-4o mean minus {j['label']} mean, all charters (pp) — "
                          "positive means GPT-4o is the more lenient judge"),
                    method=src)
            reg.put(f"{k}.model_display", JUDGE_SHORT.get(slug, j["label"]), fmt="{}",
                    desc=f"Display name for the {j['label']} cross-model judge",
                    method="fixed short label for the judge model used in this rescore; "
                           "carried as a macro so the manuscript never hand-types the "
                           "version number")

            for cond in CONDITIONS:
                c = j["by_condition"][cond]
                reg.put(f"{k}.{cond}.mean", c["judge_mean"],
                        desc=f"{CONDITION_LABELS[cond]} mean under {j['label']} (%)",
                        method=src)
                reg.put(f"{k}.{cond}.sd", c["judge_sd"],
                        desc=f"{CONDITION_LABELS[cond]} charter SD under {j['label']} (%)",
                        method=src)
                reg.put(f"{k}.{cond}.gap", c["leniency_gap"], fmt="diff",
                        desc=(f"GPT-4o minus {j['label']} on {CONDITION_LABELS[cond]} (pp)"),
                        method=src)

            for slug_c, rep in sorted(j["replication"].items()):
                jc = rep["judge"]
                cm = (src + "; Wilcoxon signed-rank on the 25 specification means, "
                            "identical to the primary inferential test")
                reg.put(f"{k}.{slug_c}.diff", jc["mean_diff"], fmt="diff",
                        desc=f"{rep['label']} under {j['label']} (pp)", method=cm)
                reg.put(f"{k}.{slug_c}.p", jc["p_signed_rank"], fmt="p",
                        desc=f"{rep['label']} signed-rank p under {j['label']}", method=cm)
                reg.put(f"{k}.{slug_c}.ci_low", jc["ci_low"], fmt="diff",
                        desc=f"{rep['label']} 95% CI lower bound under {j['label']} (pp)",
                        method=cm)
                reg.put(f"{k}.{slug_c}.ci_high", jc["ci_high"], fmt="diff",
                        desc=f"{rep['label']} 95% CI upper bound under {j['label']} (pp)",
                        method=cm)

            cl = j.get("cost_and_latency") or {}
            if cl.get("total_cost_usd") is not None:
                reg.put(f"{k}.cost", cl["total_cost_usd"], fmt="{:.2f}",
                        desc=f"Total API cost to rescore the corpus with {j['label']} (USD)",
                        method="summed from the usage block returned with each scoring call")
            if cl.get("mean_latency_s") is not None:
                reg.put(f"{k}.latency", cl["mean_latency_s"],
                        desc=f"Mean per-charter scoring latency, {j['label']} (s)",
                        method="wall-clock time per scoring call, measured client-side")

        any_judge = next(iter(judges["judges"].values()))
        reg.put("judge.primary_overall_mean", any_judge["overall_primary_mean"],
                desc="Overall charter mean under the primary GPT-4o judge (%)",
                method="mean of all 375 charter percentages in the authoritative scoring file")
        reg.put("judge.n", judges["n_judges"], fmt="int",
                desc="Independent judge models used to rescore the corpus",
                method="count of completed cross-model rescoring runs")

    # ---- Generation-side operational cost and latency (task 2.4) -------------
    if gencost:
        m = ("dedicated instrumentation run (scripts/instrument_generation.py): the three "
             "condition prompts replayed over the 25 specifications under byte-identical "
             "configuration with OpenRouter cost accounting enabled; the evaluation corpus "
             "was not regenerated")
        reg.put("gen.model_display", gencost["model_display"], fmt="{}",
                desc="Generator model, display form", method="fixed label for openai/gpt-4o")
        reg.put("gen.n_calls", gencost["n_calls"], fmt="int",
                desc="Instrumented generation calls (3 conditions x 25 specifications)", method=m)
        reg.put("gen.total_cost", gencost["total_cost_usd"], fmt="{:.2f}",
                desc="Total measured API cost of the instrumentation run (USD)", method=m)
        for cond in CONDITIONS:
            c = gencost["per_condition"][cond]
            lbl = CONDITION_LABELS[cond]
            reg.put(f"gen.{cond}.latency_mean", c["latency_mean"], fmt="{:.1f}",
                    desc=f"{lbl}: mean generation latency per specification (s)", method=m)
            reg.put(f"gen.{cond}.latency_median", c["latency_median"], fmt="{:.1f}",
                    desc=f"{lbl}: median generation latency per specification (s)", method=m)
            reg.put(f"gen.{cond}.latency_max", c["latency_max"], fmt="{:.1f}",
                    desc=f"{lbl}: slowest generation call (s)", method=m)
            reg.put(f"gen.{cond}.tok_in_mean", c["tok_in_mean"], fmt="{:.0f}",
                    desc=f"{lbl}: mean prompt tokens per specification", method=m)
            reg.put(f"gen.{cond}.tok_out_mean", c["tok_out_mean"], fmt="{:.0f}",
                    desc=f"{lbl}: mean completion tokens per specification", method=m)
            reg.put(f"gen.{cond}.tok_in_total", c["tok_in_total"], fmt="int",
                    desc=f"{lbl}: total prompt tokens over 25 specifications", method=m)
            reg.put(f"gen.{cond}.tok_out_total", c["tok_out_total"], fmt="int",
                    desc=f"{lbl}: total completion tokens over 25 specifications", method=m)
            reg.put(f"gen.{cond}.cost_total", c["cost_total"], fmt="{:.4f}",
                    desc=f"{lbl}: total measured API cost over 25 specifications (USD)", method=m)
            reg.put(f"gen.{cond}.cost_per_spec", c["cost_per_spec"], fmt="{:.4f}",
                    desc=f"{lbl}: measured API cost per specification (USD, 5 charters)", method=m)
            reg.put(f"gen.{cond}.cost_per_charter", c["cost_per_charter"], fmt="{:.4f}",
                    desc=f"{lbl}: measured API cost per charter (USD)", method=m)
            reg.put(f"gen.{cond}.cost_per_hundred", c["cost_per_spec"] * 100, fmt="{:.2f}",
                    desc=f"{lbl}: projected API cost for 100 specifications (USD)", method=m)
            reg.put(f"gen.{cond}.retries", c["total_retries"], fmt="int",
                    desc=f"{lbl}: transport retries across 25 calls", method=m)
            reg.put(f"gen.{cond}.truncated", c["truncated"], fmt="int",
                    desc=f"{lbl}: replies stopped at the 1500-token ceiling", method=m)
            reg.put(f"gen.{cond}.not_five", c["not_five_charters"], fmt="int",
                    desc=f"{lbl}: calls whose output did not yield exactly 5 charters", method=m)
        reg.put("gen.etcg.json_invalid", gencost["per_condition"]["etcg"]["json_invalid"], fmt="int",
                desc="ETCG: responses that failed JSON parsing", method=m)
        reg.put("gen.etcg.schema_incomplete",
                gencost["per_condition"]["etcg"]["schema_incomplete"], fmt="int",
                desc="ETCG: valid-JSON responses missing one or more required charter fields",
                method=m)
        for cond in ("intermediate", "baseline"):
            reg.put(f"gen.{cond}.underextracted",
                    gencost["per_condition"][cond]["underextracted"], fmt="int",
                    desc=f"{CONDITION_LABELS[cond]}: free-text replies yielding fewer than 5 "
                         f"parseable charters", method=m)

    # ---- Multi-model generalisation (task 2.3) ------------------------------
    if multimodel and multimodel.get("models"):
        m = ("the three conditions regenerated on this model family and scored by the "
             "fixed GPT-4o rubric judge; contrasts recomputed at the specification level "
             "by Wilcoxon signed-rank, identically to the primary analysis")
        reg.put("mm.n_models", multimodel["n_models"], fmt="int",
                desc="Additional generator model families evaluated (task 2.3)",
                method="count of completed multi-model score sets")
        reg.put("mm.n_generators", multimodel["n_generators"], fmt="int",
                desc="Generator model families in total, including the primary GPT-4o run",
                method="additional generators + 1")
        reg.put("mm.role_positive_in", multimodel["role_positive_in"], fmt="int",
                desc="Generators where role framing raises the mean (Baseline to Intermediate)",
                method=m + "; count over all generators including GPT-4o")
        reg.put("mm.role_significant_in", multimodel["role_significant_in"], fmt="int",
                desc="Generators where the role-framing contrast reaches p < 0.05",
                method=m + "; count over all generators including GPT-4o")
        reg.put("mm.schema_nonsignificant_in", multimodel["schema_nonsignificant_in"], fmt="int",
                desc="Generators where the output-schema contrast does not reach p < 0.05",
                method=m + "; count over all generators including GPT-4o")
        reg.put("mm.schema_negative_in", multimodel["schema_negative_in"], fmt="int",
                desc="Generators where adding the output schema lowers the mean",
                method=m + "; count over all generators including GPT-4o")
        reg.put("mm.baseline_last_in", multimodel["baseline_last_in"], fmt="int",
                desc="Generators whose lowest-scoring condition is Baseline",
                method=m + "; count over all generators including GPT-4o")
        if multimodel.get("total_cost_usd") is not None:
            reg.put("mm.total_cost", multimodel["total_cost_usd"], fmt="{:.2f}",
                    desc="Total API cost to regenerate and rescore all three conditions for the additional generator families (USD)",
                    method="sum of completed per-call generation instrumentation and per-model scoring cost metadata")
        for slug, md in sorted(multimodel["models"].items()):
            k = f"mm.{slug}"
            reg.put(f"{k}.short", md["short"], fmt="{}",
                    desc=f"Display name for the {md['label']} generator",
                    method="fixed short label")
            reg.put(f"{k}.n_charters", md["n_charters"], fmt="int",
                    desc=f"{md['short']}: charters scored", method=m)
            for cond in CONDITIONS:
                reg.put(f"{k}.{cond}_mean", md["condition"][cond]["mean"], fmt="{:.2f}",
                        desc=f"{md['short']}: {CONDITION_LABELS[cond]} mean quality (%)", method=m)
                reg.put(f"{k}.{cond}_sd", md["condition"][cond]["sd"], fmt="{:.2f}",
                        desc=f"{md['short']}: {CONDITION_LABELS[cond]} SD across specification means (%)",
                        method=m)
            reg.put(f"{k}.charters_per_cell_min", md["condition"]["baseline"]["min_charters_per_cell"],
                    fmt="int",
                    desc=f"{md['short']}: fewest charters the splitter recovered from any "
                         f"free-text cell", method=m)
            for slug_c, res in md["contrasts"].items():
                reg.put(f"{k}.{slug_c}.diff", res["mean_diff"], fmt="diff",
                        desc=f"{md['short']}: {res['label']} — specification-level mean difference (pp)",
                        method=m)
                reg.put(f"{k}.{slug_c}.p", res["p_signed_rank"], fmt="p",
                        desc=f"{md['short']}: {res['label']} — Wilcoxon signed-rank p", method=m)
                reg.put(f"{k}.{slug_c}.ci_low", res["ci_low"], fmt="diff",
                        desc=f"{md['short']}: {res['label']} — 95% CI lower bound (pp)", method=m)
                reg.put(f"{k}.{slug_c}.ci_high", res["ci_high"], fmt="diff",
                        desc=f"{md['short']}: {res['label']} — 95% CI upper bound (pp)", method=m)
            reg.put(f"{k}.etcg_vs_primary_etcg", md["etcg_minus_primary_etcg_pp"], fmt="diff",
                    desc=f"{md['short']}: ETCG mean minus the primary GPT-4o ETCG mean (pp)",
                    method=m)


# ── Emission ─────────────────────────────────────────────────────────────────


def emit_macros(reg: Registry, path: Path) -> dict[str, str]:
    lines = [
        "% macros.tex — GENERATED BY scripts/analysis.py. DO NOT EDIT.",
        "%",
        "% Every numeric value in the manuscript prose must come from one of these macros.",
        "% Hand-typing a number into main.tex is what produced the round 1 inconsistencies.",
        f"% Generated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}",
        "",
    ]
    mapping = {}
    seen = {}
    for key, entry in reg.entries():
        macro = key_to_macro(key)
        if macro in seen:
            raise ValueError(f"macro collision: {key} and {seen[macro]} both -> {macro}")
        seen[macro] = key
        mapping[key] = macro
        lines.append(f"% {entry['desc']}")
        lines.append(f"\\newcommand{{\\{macro}}}{{{render(entry['value'], entry['fmt'])}}}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")
    return mapping


def tex_table(caption: str, label: str, colspec: str, header: list[str],
              rows: list[list[str]], notes: list[str] | None = None) -> str:
    out = [
        "% GENERATED BY scripts/analysis.py. DO NOT EDIT.",
        "\\begin{table}[t]",
        f"\\caption{{{caption}}}",
        f"\\label{{{label}}}",
        f"\\begin{{tabular}}{{@{{}}{colspec}@{{}}}}",
        "\\toprule",
        " & ".join(header) + " \\\\",
        "\\midrule",
    ]
    out.extend(" & ".join(r) + " \\\\" for r in rows)
    if notes:
        out.append("\\addlinespace[0.25em]")
        ncols = len(header)
        for note in notes:
            out.append(f"\\multicolumn{{{ncols}}}{{@{{}}p{{\\linewidth}}@{{}}}}{{\\footnotesize\\raggedright {note}}} \\\\")
    out.append("\\bottomrule")
    out.extend(["\\end{tabular}", "\\end{table}", ""])
    return "\n".join(out)


def emit_tables(reg: Registry, out_dir: Path) -> list[Path]:
    written = []
    r = reg.rendered

    # Table: RQ1 overall quality by condition.
    rows = []
    for c in CONDITIONS:
        rows.append([
            CONDITION_LABELS[c],
            r(f"rq1.{c}.mean"),
            f"[{r(f'rq1.{c}.ci_low')},\\ {r(f'rq1.{c}.ci_high')}]",
            r(f"rq1.{c}.sd_spec"),
            r(f"rq1.{c}.sd_charter"),
            f"{r(f'rq1.{c}.ceiling_n')}/125 ({r(f'rq1.{c}.ceiling_pct')}\\%)",
        ])
    # Rendered values may themselves contain math ($-$, $<$0.001), so symbols are set in
    # their own math group and the value follows outside it. Nesting them breaks the build.
    notes = [f"Pairwise Wilcoxon signed-rank on {r('corpus.n_specs')} paired specification means:"]
    for _, _, slug, label in CONTRASTS:
        notes.append(
            f"{label}: {r(f'rq1.{slug}.mean_diff')}pp, "
            f"95\\% CI [{r(f'rq1.{slug}.ci_low')}, {r(f'rq1.{slug}.ci_high')}], "
            f"$W$ = {r(f'rq1.{slug}.w')}, $p$ = {r(f'rq1.{slug}.p')}, "
            f"$d_z$ = {r(f'rq1.{slug}.dz')}"
        )
    if "mixed.icc" in reg.keys():
        estimates = "; ".join(
            f"{label.split(' (')[0]} {r(f'mixed.{slug}.coef')}pp, $p$ = {r(f'mixed.{slug}.p')}"
            for _, _, slug, label in CONTRASTS
        )
        notes.append("Mixed-effects estimates (random intercept per specification, "
                     f"ICC = {r('mixed.icc')}): {estimates}.")
    path = out_dir / "table-rq1-overall.tex"
    path.write_text(tex_table(
        caption=(f"Overall charter quality by condition. Means and confidence intervals are "
                 f"computed at the specification level ($n = {r('corpus.n_specs')}$ "
                 f"specifications per condition); charter-level SD is reported as a "
                 f"descriptive measure of output consistency."),
        label="tab:rq1", colspec="lrrrrr",
        header=["Condition", "Mean (\\%)", "95\\% CI (\\%)", "SD$_{\\mathrm{spec}}$ (\\%)",
                "SD$_{\\mathrm{charter}}$ (\\%)", "At ceiling"],
        rows=rows, notes=notes), encoding="utf-8")
    written.append(path)

    # Table: RQ2 per-dimension means and contrasts.
    rows = []
    for d in DIMENSIONS:
        rows.append([
            DIMENSION_LABELS[d],
            r(f"rq2.{d}.etcg.mean"), r(f"rq2.{d}.intermediate.mean"), r(f"rq2.{d}.baseline.mean"),
            r(f"rq2.{d}.full_framework.p"), r(f"rq2.{d}.role_guidance.p"), r(f"rq2.{d}.schema.p"),
            r(f"rq2.{d}.full_framework.p_holm"), r(f"rq2.{d}.role_guidance.p_holm"),
            r(f"rq2.{d}.schema.p_holm"),
        ])
    path = out_dir / "table-rq2-dimensions.tex"
    path.write_text(tex_table(
        caption=(f"Mean dimension scores by condition (scale 1--3), computed at the "
                 f"specification level. Raw and Holm-adjusted $p$-values from Wilcoxon "
                 f"signed-rank tests across the family of {r('rq2.n_tests')} dimension-level "
                 f"comparisons."),
        label="tab:rq2", colspec="lrrrrrrrrr",
        header=["Dimension", "ETCG", "Inter.", "Base.",
                "$p$(E--B)", "$p$(I--B)", "$p$(E--I)",
                "$p_{\\mathrm{H}}$(E--B)", "$p_{\\mathrm{H}}$(I--B)", "$p_{\\mathrm{H}}$(E--I)"],
        rows=rows,
        notes=["E=ETCG, I=Intermediate, B=Baseline. $p_{\\mathrm{H}}$ = Holm-adjusted."]),
        encoding="utf-8")
    written.append(path)

    # Table: RQ3 inter-rater reliability.
    if "rq3.n_charters" in reg.keys():
        rows = []
        for d in DIMENSIONS:
            rows.append([
                DIMENSION_LABELS[d],
                r(f"rq3.{d}.auto_mean"), r(f"rq3.{d}.human_mean"), r(f"rq3.{d}.delta"),
                r(f"rq3.{d}.exact_pct") + "\\%", r(f"rq3.{d}.within1_pct") + "\\%",
                r(f"rq3.{d}.two_point_n"),
                r(f"rq3.{d}.alpha_ordinal"), r(f"rq3.{d}.ac2"),
            ])
        rows.append(["\\midrule \\textbf{Overall}",
                     r("rq3.auto_mean_pct") + "\\%", r("rq3.human_mean_pct") + "\\%",
                     r("rq3.mean_delta_pct"),
                     r("rq3.exact_pct") + "\\%", r("rq3.within1_pct") + "\\%",
                     r("rq3.two_point_n"), r("rq3.alpha"), r("rq3.ac2")])
        path = out_dir / "table-rq3-irr.tex"
        path.write_text(tex_table(
            caption=(f"Inter-rater reliability: automated evaluator versus independent human "
                     f"reviewer ($n = {r('rq3.n_charters')}$ charters, "
                     f"{r('rq3.n_ratings')} dimension ratings)."),
            label="tab:rq3", colspec="lrrrrrrrr",
            header=["Dimension", "Auto", "Human", "$\\Delta$", "Exact", "Within-1",
                    "2-pt", "$\\alpha$", "AC2"],
            rows=rows,
            notes=[f"Mean absolute difference in total score: {r('rq3.mean_abs_total_diff')} "
                   f"points of {MAX_CHARTER_POINTS}.",
                   f"2-pt = ratings differing by two scale points ({r('rq3.two_point_n')} of "
                   f"{r('rq3.n_ratings')}, {r('rq3.two_point_pct')}\\%).",
                   f"$\\alpha$ = Krippendorff's alpha (ordinal); AC2 = Gwet's AC2 (linear weights)."]),
            encoding="utf-8")
        written.append(path)

    # Table: RQ3 round 2 — automated evaluator versus the second human rater.
    if "rq3.r2.n_charters" in reg.keys():
        rows = []
        for d in DIMENSIONS:
            rows.append([
                DIMENSION_LABELS[d],
                r(f"rq3.r2.{d}.auto_mean"), r(f"rq3.r2.{d}.human_mean"),
                r(f"rq3.r2.{d}.delta"),
                r(f"rq3.r2.{d}.exact_pct") + "\\%", r(f"rq3.r2.{d}.within1_pct") + "\\%",
                r(f"rq3.r2.{d}.two_point_n"),
                r(f"rq3.r2.{d}.alpha_ordinal"), r(f"rq3.r2.{d}.ac2"),
            ])
        rows.append(["\\midrule \\textbf{Overall}",
                     r("rq3.r2.auto_mean_pct") + "\\%", r("rq3.r2.human_mean_pct") + "\\%",
                     r("rq3.r2.mean_delta_pct"),
                     r("rq3.r2.exact_pct") + "\\%", r("rq3.r2.within1_pct") + "\\%",
                     r("rq3.r2.two_point_n"), r("rq3.r2.alpha"), r("rq3.r2.ac2")])
        path = out_dir / "table-rq3-rater2.tex"
        path.write_text(tex_table(
            caption=(f"Round 2 inter-rater reliability: automated evaluator versus a second "
                     f"independent human reviewer, across all three conditions "
                     f"($n = {r('rq3.r2.n_charters')}$ charters, "
                     f"{r('rq3.r2.n_ratings')} dimension ratings)."),
            label="tab:rq3r2", colspec="lrrrrrrrr",
            header=["Dimension", "Auto", "Rater 2", "$\\Delta$", "Exact", "Within-1",
                    "2-pt", "$\\alpha$", "AC2"],
            rows=rows,
            notes=[f"Mean absolute difference in total score: "
                   f"{r('rq3.r2.mean_abs_total_diff')} points of {MAX_CHARTER_POINTS}.",
                   f"2-pt = ratings differing by two scale points "
                   f"({r('rq3.r2.two_point_n')} of {r('rq3.r2.n_ratings')}, "
                   f"{r('rq3.r2.two_point_pct')}\\%).",
                   f"$\\Delta$ = rater 2 mean minus automated mean.",
                   f"$\\alpha$ = Krippendorff's alpha (ordinal); AC2 = Gwet's AC2 "
                   f"(linear weights)."]),
            encoding="utf-8")
        written.append(path)

    # Table: RQ3 round 2 — automated vs rater 2, split by condition.
    r2_conds = [c for c in CONDITIONS if f"rq3.r2.{c}.n_charters" in reg.keys()]
    if r2_conds:
        rows = []
        for c in r2_conds:
            rows.append([
                CONDITION_LABELS[c],
                r(f"rq3.r2.{c}.n_charters"),
                r(f"rq3.r2.{c}.auto_mean_pct") + "\\%",
                r(f"rq3.r2.{c}.human_mean_pct") + "\\%",
                r(f"rq3.r2.{c}.mean_delta_pct"),
                r(f"rq3.r2.{c}.exact_pct") + "\\%",
                r(f"rq3.r2.{c}.within1_pct") + "\\%",
                r(f"rq3.r2.{c}.two_point_n"),
                r(f"rq3.r2.{c}.ac2"),
            ])
        notes = ["$\\Delta$ = rater 2 mean minus automated mean, at the charter level.",
                 "AC2 = Gwet's AC2 (linear weights)."]
        if "rq3.cons.n_charters" in reg.keys():
            notes.append(
                f"Against the mean of both human raters on the "
                f"{r('rq3.cons.n_charters')}-charter two-rater overlap, the automated "
                f"evaluator sits {r('rq3.cons.mean_delta_pct')} pp from consensus "
                f"(mean absolute difference {r('rq3.cons.mean_abs_total_diff')} of "
                f"{MAX_CHARTER_POINTS} points).")
        path = out_dir / "table-rq3-rater2-bycond.tex"
        path.write_text(tex_table(
            caption=("Round 2 inter-rater reliability by condition: automated evaluator "
                     "versus the second human reviewer. The Intermediate row is the first "
                     "human validation of the schema-ablation condition."),
            label="tab:rq3r2cond", colspec="lrrrrrrrr",
            header=["Condition", "$n$", "Auto", "Rater 2", "$\\Delta$", "Exact",
                    "Within-1", "2-pt", "AC2"],
            rows=rows, notes=notes),
            encoding="utf-8")
        written.append(path)

    # Table: RQ3 human--human agreement (rater 1 versus rater 2 on the overlap).
    if "rq3.hh.n_charters" in reg.keys():
        rows = []
        for d in DIMENSIONS:
            rows.append([
                DIMENSION_LABELS[d],
                r(f"rq3.hh.{d}.auto_mean"), r(f"rq3.hh.{d}.human_mean"),
                r(f"rq3.hh.{d}.delta"),
                r(f"rq3.hh.{d}.exact_pct") + "\\%", r(f"rq3.hh.{d}.within1_pct") + "\\%",
                r(f"rq3.hh.{d}.two_point_n"),
                r(f"rq3.hh.{d}.alpha_ordinal"), r(f"rq3.hh.{d}.ac2"),
            ])
        rows.append(["\\midrule \\textbf{Overall}",
                     r("rq3.hh.r1_mean_pct") + "\\%", r("rq3.hh.r2_mean_pct") + "\\%",
                     r("rq3.hh.mean_delta_pct"),
                     r("rq3.hh.exact_pct") + "\\%", r("rq3.hh.within1_pct") + "\\%",
                     r("rq3.hh.two_point_n"), r("rq3.hh.alpha"), r("rq3.hh.ac2")])
        path = out_dir / "table-rq3-human-human.tex"
        path.write_text(tex_table(
            caption=(f"Human--human agreement: the two independent human reviewers on the "
                     f"{r('rq3.hh.n_charters')}-charter overlap they both scored "
                     f"({r('rq3.hh.n_ratings')} dimension ratings). Separates rubric "
                     f"ambiguity from single-rater idiosyncrasy."),
            label="tab:rq3hh", colspec="lrrrrrrrr",
            header=["Dimension", "Rater 1", "Rater 2", "$\\Delta$", "Exact", "Within-1",
                    "2-pt", "$\\alpha$", "AC2"],
            rows=rows,
            notes=[f"Mean absolute difference in total score: "
                   f"{r('rq3.hh.mean_abs_total_diff')} points of {MAX_CHARTER_POINTS}.",
                   f"$\\Delta$ = rater 2 mean minus rater 1 mean.",
                   f"2-pt = ratings differing by two scale points ({r('rq3.hh.two_point_n')} "
                   f"of {r('rq3.hh.n_ratings')}, {r('rq3.hh.two_point_pct')}\\%).",
                   f"$\\alpha$ = Krippendorff's alpha (ordinal); AC2 = Gwet's AC2 "
                   f"(linear weights)."]),
            encoding="utf-8")
        written.append(path)

    # Table: RQ4 specification richness.
    rows = []
    for c in CONDITIONS:
        rows.append([
            CONDITION_LABELS[c],
            r(f"rq4.{c}.sparse.mean"), r(f"rq4.{c}.sparse.sd"),
            r(f"rq4.{c}.structured.mean"), r(f"rq4.{c}.structured.sd"),
            r(f"rq4.{c}.contrast_diff"), r(f"rq4.{c}.contrast_d"), r(f"rq4.{c}.contrast_p_holm"),
        ])
    path = out_dir / "table-rq4-richness.tex"
    path.write_text(tex_table(
        caption=(f"Charter quality by specification richness, at the specification level "
                 f"(sparse $n = {r('rq4.etcg.sparse.n')}$, "
                 f"structured $n = {r('rq4.etcg.structured.n')}$)."),
        label="tab:rq4", colspec="lrrrrrrr",
        header=["Condition", "Sparse M", "Sparse SD", "Struct.\\ M", "Struct.\\ SD",
                "$\\Delta$", "$d$", "$p$"],
        rows=rows,
        notes=["Structured and sparse specifications are independent groups, so an unpaired "
               "Mann--Whitney test is used for each contrast. $p$ is Holm-adjusted across "
               f"the {r('rq4.n_tests')} condition-wise RQ4 contrasts."]),
        encoding="utf-8")
    written.append(path)

    # Table: RQ4 dimension breakdown by richness.
    rows = []
    for d in DIMENSIONS:
        rows.append([DIMENSION_LABELS[d], r(f"rq4.dim.{d}.sparse"),
                     r(f"rq4.dim.{d}.structured"), r(f"rq4.dim.{d}.delta")])
    path = out_dir / "table-rq4-dimensions.tex"
    path.write_text(tex_table(
        caption=(f"ETCG dimension means by specification richness, at the specification level "
                 f"(sparse $n = {r('rq4.etcg.sparse.n')}$, "
                 f"structured $n = {r('rq4.etcg.structured.n')}$)."),
        label="tab:rq4dims", colspec="lrrr",
        header=["Dimension", "Sparse", "Structured", "$\\Delta$"],
        rows=rows), encoding="utf-8")
    written.append(path)

    # Table: cross-model judge agreement and severity.
    judge_slugs = sorted(
        k.split(".")[1] for k in reg.keys()
        if k.startswith("judge.") and k.endswith(".overall_mean")
    )
    if judge_slugs:
        rows = [[
            "GPT-4o (primary)",
            r("judge.primary_overall_mean"),
            "---", "---", "---", "---",
        ]]
        for slug in judge_slugs:
            rows.append([
                JUDGE_SHORT.get(slug, slug),
                r(f"judge.{slug}.overall_mean"),
                r(f"judge.{slug}.leniency_gap"),
                r(f"judge.{slug}.exact"),
                r(f"judge.{slug}.within_one"),
                r(f"judge.{slug}.ac2"),
            ])
        path = out_dir / "table-judge-agreement.tex"
        path.write_text(tex_table(
            caption=("Cross-model judge agreement with the primary GPT-4o evaluator. "
                     "Every judge scored the identical frozen corpus with the identical "
                     "rubric prompt at temperature 0; only the judge model differs. "
                     "$\\Delta$ is the primary mean minus the judge mean, so a positive "
                     "value indicates an overall judge-severity gap: GPT-4o assigned "
                     "higher scores under the shared rubric, not a preference for a "
                     "particular generator."),
            label="tab:judgeagreement", colspec="lrrrrr",
            header=["Judge", "Mean (\\%)", "$\\Delta$ (pp)", "Exact (\\%)",
                    "Within-one (\\%)", "AC2"],
            rows=rows,
            notes=[f"Agreement computed over {r(f'judge.{judge_slugs[0]}.n_ratings')} "
                   "dimension ratings (375 charters $\\times$ 5 dimensions)."]),
            encoding="utf-8")
        written.append(path)

        # Table: does each contrast survive a change of judge?
        header = ["Contrast", "GPT-4o $\\Delta$", "$p$"]
        for slug in judge_slugs:
            header += [f"{JUDGE_SHORT.get(slug, slug)} $\\Delta$", "$p$"]
        rows = []
        for _, _, slug_c, label in CONTRASTS:
            row = [label.split(" (")[0],
                   r(f"rq1.{slug_c}.mean_diff"), r(f"rq1.{slug_c}.p")]
            for slug in judge_slugs:
                row += [r(f"judge.{slug}.{slug_c}.diff"), r(f"judge.{slug}.{slug_c}.p")]
            rows.append(row)
        path = out_dir / "table-judge-replication.tex"
        path.write_text(tex_table(
            caption=("Each pairwise contrast recomputed under every judge. Differences are "
                     "in percentage points between paired specification means "
                     f"($n = {r('corpus.n_specs')}$), tested by Wilcoxon signed-rank, "
                     "identically to the primary analysis."),
            label="tab:judgereplication", colspec="l" + "rr" * (1 + len(judge_slugs)),
            header=header, rows=rows),
            encoding="utf-8")
        written.append(path)

    # Table: generation-side operational cost and latency (task 2.4).
    if "gen.n_calls" in reg.keys():
        rows = []
        for c in CONDITIONS:
            rows.append([
                CONDITION_LABELS[c],
                r(f"gen.{c}.latency_mean"),
                r(f"gen.{c}.tok_in_mean"),
                r(f"gen.{c}.tok_out_mean"),
                r(f"gen.{c}.cost_per_spec"),
                r(f"gen.{c}.cost_per_hundred"),
            ])
        notes = [
            f"Dedicated instrumentation run ({r('gen.model_display')}, identical prompts, "
            "temperature 0.2, 1500-token ceiling); the evaluation corpus was not "
            f"regenerated. {r('corpus.n_specs')} calls per condition, one per specification, "
            "five charters each.",
            f"Transport retries: ETCG {r('gen.etcg.retries')}, Intermediate "
            f"{r('gen.intermediate.retries')}, Baseline {r('gen.baseline.retries')}. "
            f"Replies stopped at the token ceiling: ETCG {r('gen.etcg.truncated')}, "
            f"Intermediate {r('gen.intermediate.truncated')}, Baseline "
            f"{r('gen.baseline.truncated')}.",
            f"ETCG JSON parse failures: {r('gen.etcg.json_invalid')}/{r('corpus.n_specs')}; "
            f"schema-incomplete: {r('gen.etcg.schema_incomplete')}/{r('corpus.n_specs')}. "
            f"Free-text replies yielding fewer than five charters: Intermediate "
            f"{r('gen.intermediate.underextracted')}/{r('corpus.n_specs')}, Baseline "
            f"{r('gen.baseline.underextracted')}/{r('corpus.n_specs')}.",
        ]
        path = out_dir / "table-generation-cost.tex"
        path.write_text(tex_table(
            caption=("Operational cost and latency of charter generation, by condition. "
                     "Latency is wall-clock time for the single API call that produces all "
                     "five charters for a specification; token counts and USD cost are read "
                     "back from the provider's usage accounting."),
            label="tab:gencost", colspec="lrrrrr",
            header=["Condition", "Latency (s)", "Tok.\\ in", "Tok.\\ out",
                    "USD/spec", "USD/100 specs"],
            rows=rows, notes=notes), encoding="utf-8")
        written.append(path)

    # Table: multi-model generalisation of the ablation contrasts (task 2.3).
    mm_slugs = sorted(
        k.split(".")[1] for k in reg.keys()
        if k.startswith("mm.") and k.endswith(".etcg_mean")
    )
    if mm_slugs:
        rows = [[
            "GPT-4o (primary)",
            r("rq1.baseline.mean"), r("rq1.intermediate.mean"), r("rq1.etcg.mean"),
            r("rq1.role_guidance.mean_diff"), r("rq1.role_guidance.p"),
            r("rq1.schema.mean_diff"), r("rq1.schema.p"),
        ]]
        for slug in mm_slugs:
            rows.append([
                r(f"mm.{slug}.short"),
                r(f"mm.{slug}.baseline_mean"), r(f"mm.{slug}.intermediate_mean"),
                r(f"mm.{slug}.etcg_mean"),
                r(f"mm.{slug}.role_guidance.diff"), r(f"mm.{slug}.role_guidance.p"),
                r(f"mm.{slug}.schema.diff"), r(f"mm.{slug}.schema.p"),
            ])
        notes = [
            "Every generator's charters were scored by the same fixed GPT-4o rubric judge "
            "(temperature 0); only the model that wrote the charters differs. Means are over "
            f"{r('corpus.n_specs')} specification-level means per condition; $\\Delta$ columns "
            "are paired specification-mean differences (pp), Wilcoxon signed-rank.",
            f"Role framing (Baseline $\\rightarrow$ Intermediate) raises the mean for "
            f"{r('mm.role_positive_in')}/{r('mm.n_generators')} generators. Adding the output "
            f"schema (Intermediate $\\rightarrow$ ETCG) does not reach significance for "
            f"{r('mm.schema_nonsignificant_in')}/{r('mm.n_generators')} and lowers the mean for "
            f"{r('mm.schema_negative_in')}/{r('mm.n_generators')}. Baseline is the "
            f"lowest-scoring condition for {r('mm.baseline_last_in')}/{r('mm.n_generators')}.",
        ]
        path = out_dir / "table-multimodel.tex"
        path.write_text(tex_table(
            caption=("The ablation contrasts recomputed with the charter generator swapped "
                     "for other model families. The prompt design, the conditions, the "
                     "specifications, and the GPT-4o rubric judge are all held fixed."),
            label="tab:multimodel", colspec="lrrrrrrr",
            header=["Generator", "Base.", "Inter.", "ETCG",
                    "role $\\Delta$", "$p$", "schema $\\Delta$", "$p$"],
            rows=rows, notes=notes), encoding="utf-8")
        written.append(path)

    return written


def emit_traceability(reg: Registry, mapping: dict[str, str], path: Path,
                      inputs: dict, tables: list[Path], irr: dict | None,
                      irr_r2: dict | None = None) -> None:
    lines = [
        "# Traceability appendix",
        "",
        "Every numeric value reported in the manuscript is produced by "
        "`scripts/analysis.py` and listed below with the computation that produced it.",
        "",
        "Prose cites the LaTeX macro; tables are `\\input` from generated files. No value "
        "is typed by hand into `main.tex`.",
        "",
        f"Generated: {datetime.now(timezone.utc).isoformat(timespec='seconds')}",
        "",
        "## Inputs",
        "",
    ]
    for name, meta in inputs.items():
        lines.append(f"- **{name}** — `{meta['path']}`  ")
        lines.append(f"  SHA-256 `{meta['sha256']}` · {meta['bytes']:,} bytes")
    lines.extend(["", "## Generated tables", ""])
    for t in tables:
        lines.append(f"- `{t.name}`")

    if irr:
        lines.extend([
            "", "## Two-point disagreements in the IRR sample", "",
            "The round 1 submission stated that no two-point disagreements occurred. "
            f"There are **{irr['overall']['two_point_n']}**, listed here in full.", "",
            "| Charter | Spec | Condition | Dimension | Automated | Human |",
            "|---|---|---|---|---|---|",
        ])
        for loc in irr["overall"]["two_point_locations"]:
            lines.append(
                f"| {loc['neutral_id']} | {loc['spec_id']} | {loc['condition']} | "
                f"{DIMENSION_LABELS[loc['dimension']]} | {loc['auto']} | {loc['human']} |")
        missing = irr["overall"]["conditions_missing"]
        if missing:
            note = (f"> **Coverage gap (round 1).** The round 1 IRR sample covers "
                    f"{', '.join(irr['overall']['conditions_covered'])} but not "
                    f"{', '.join(missing)}.")
            if irr_r2 and irr_r2.get("by_condition"):
                note += (" Round 2 closes this: a second independent rater scored all "
                         "three conditions, including "
                         f"{', '.join(sorted(irr_r2['by_condition']))}.")
            else:
                note += (" The schema ablation contrast is therefore not independently "
                         "validated. Task 3.1 addresses this.")
            lines.extend(["", note, ""])

    if irr_r2:
        allr = irr_r2["all"]["overall"]
        lines.extend([
            "", "## Round 2 — second independent human rater", "",
            f"Automated evaluator versus rater 2 across **{allr['n_charters']} charters** "
            f"({', '.join(allr['conditions_covered'])}): exact "
            f"{allr['exact_pct']:.1f}%, within-one {allr['within1_pct']:.1f}%, "
            f"**{allr['two_point_n']}** two-point disagreements, "
            f"Gwet's AC2 {allr['ac2']:.3f}. Rater 2 mean "
            f"{allr['human_mean_pct']:.1f}% versus automated "
            f"{allr['auto_mean_pct']:.1f}% "
            f"({allr['human_mean_pct'] - allr['auto_mean_pct']:+.1f} pp).", "",
            "### Two-point disagreements, automated vs rater 2 (all conditions)", "",
            "| Charter | Spec | Condition | Dimension | Automated | Rater 2 |",
            "|---|---|---|---|---|---|",
        ])
        for loc in allr["two_point_locations"]:
            lines.append(
                f"| {loc['neutral_id']} | {loc['spec_id']} | {loc['condition']} | "
                f"{DIMENSION_LABELS[loc['dimension']]} | {loc['auto']} | {loc['human']} |")
        if "human_human" in irr_r2:
            hh = irr_r2["human_human"]["overall"]
            lines.extend([
                "", "### Human--human agreement (rater 1 vs rater 2, overlap)", "",
                f"On the **{hh['n_charters']}-charter overlap** both humans scored: exact "
                f"{hh['exact_pct']:.1f}%, within-one {hh['within1_pct']:.1f}%, "
                f"**{hh['two_point_n']}** two-point disagreements, AC2 {hh['ac2']:.3f}, "
                f"Krippendorff's alpha {hh['alpha_ordinal']:.3f}. Rater 1 mean "
                f"{hh['auto_mean_pct']:.1f}%, rater 2 mean {hh['human_mean_pct']:.1f}% "
                f"({hh['human_mean_pct'] - hh['auto_mean_pct']:+.1f} pp).", "",
                "| Charter | Spec | Condition | Dimension | Rater 1 | Rater 2 |",
                "|---|---|---|---|---|---|",
            ])
            for loc in hh["two_point_locations"]:
                lines.append(
                    f"| {loc['neutral_id']} | {loc['spec_id']} | {loc['condition']} | "
                    f"{DIMENSION_LABELS[loc['dimension']]} | {loc['auto']} | {loc['human']} |")
        if "consensus" in irr_r2:
            co = irr_r2["consensus"]["overall"]
            lines.extend([
                "", f"**Automated vs two-rater consensus** on the {co['n_charters']}-charter "
                f"overlap: automated {co['auto_mean_pct']:.1f}%, consensus "
                f"{co['consensus_mean_pct']:.1f}% "
                f"({co['delta_pct']:+.1f} pp; mean absolute total difference "
                f"{co['mean_abs_total_diff']:.2f} of {MAX_CHARTER_POINTS}).", ""])

    lines.extend(["", "## Value index", "",
                  "| Key | Macro | Value | Description | Method |",
                  "|---|---|---|---|---|"])
    for key, entry in reg.entries():
        val = render(entry["value"], entry["fmt"]).replace("$<$", "&lt;").replace("|", "\\|")
        desc = entry["desc"].replace("|", "\\|")
        method = entry["method"].replace("|", "\\|")
        lines.append(f"| `{key}` | `\\{mapping[key]}` | {val} | {desc} | {method} |")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def audit_tex(reg: Registry, tex_path: Path) -> list[str]:
    """Report decimal literals in the manuscript that this script did not produce.

    Anything listed here is a hand-typed number: either it needs replacing with a macro,
    or it disagrees with the data and is exactly the class of error round 1 flagged.
    """
    text = tex_path.read_text(encoding="utf-8")
    # Strip comments before scanning.
    text = re.sub(r"(?<!\\)%.*", "", text)
    # DOIs, URLs and version strings carry decimals that are not results. Blank them out so
    # the audit reports only numbers that make an empirical claim.
    text = re.sub(r"\\url\{[^}]*\}", " ", text)
    text = re.sub(r"10\.\d{4,9}/[^\s}]+", " ", text)
    # Layout lengths and licence names are not empirical claims either.
    text = re.sub(r"width\s*=\s*[\d.]+\\linewidth", " ", text)
    text = re.sub(r"\[[^\]]*(?:width|scale|height)\s*=\s*[\d.]+[^\]]*\]", " ", text)
    text = re.sub(r"CC[- ]BY[~ ]?\d+\.\d+", " ", text)

    produced = set()
    for _, entry in reg.entries():
        s = render(entry["value"], entry["fmt"])
        produced.add(s.replace("$<$", "<"))
        if isinstance(entry["value"], (int, float)):
            for fmt in ("{:.0f}", "{:.1f}", "{:.2f}", "{:.3f}"):
                produced.add(fmt.format(entry["value"]))

    findings = []
    for m in re.finditer(r"(?<![\w.])(\d+\.\d+)(?![\w])", text):
        literal = m.group(1)
        if literal in produced:
            continue
        line_no = text.count("\n", 0, m.start()) + 1
        context = text[max(0, m.start() - 60):m.end() + 40].replace("\n", " ").strip()
        findings.append(f"line {line_no}: {literal}  …{context}…")
    return findings


def file_meta(path: Path, root: Path | None = None) -> dict:
    """Hash an input file. Paths are recorded relative to the repository so the
    traceability appendix is portable for anyone replicating the analysis."""
    import hashlib
    data = path.read_bytes()
    shown = str(path)
    if root:
        try:
            shown = str(path.relative_to(root))
        except ValueError:
            shown = str(Path("..") / path.relative_to(root.parent))
    return {"path": shown, "sha256": hashlib.sha256(data).hexdigest(), "bytes": len(data)}


# ── Main ─────────────────────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--repo-root", type=Path, default=Path(__file__).resolve().parent.parent,
                    help="root of the supplementary repository")
    ap.add_argument("--paper-root", type=Path, default=None,
                    help="paper directory to receive generated .tex (default: ../paper-2-EMSE)")
    ap.add_argument("--no-figures", action="store_true", help="skip figure regeneration")
    ap.add_argument("--audit-tex", type=Path, default=None,
                    help="scan a .tex file for hand-typed numbers this script did not produce")
    args = ap.parse_args()

    repo = args.repo_root.resolve()
    paper = (args.paper_root or repo.parent / "paper-2-EMSE").resolve()
    results_dir = repo / "results"
    results_dir.mkdir(exist_ok=True)

    scores_path = repo / "data" / "etcg-scores.json"
    if not scores_path.exists():
        print(f"ERROR: scoring file not found: {scores_path}", file=sys.stderr)
        return 1

    prerepair_path = repo / "data" / PREREPAIR_SCORES_NAME
    if not prerepair_path.exists():
        print(f"ERROR: pre-repair scoring file not found: {prerepair_path}\n"
              f"  Required input: the manuscript's Threats section reports what "
              f"repair_baseline_split.py\n  changed, recomputed from this file. It is not "
              f"a disposable backup.", file=sys.stderr)
        return 1

    print(f"reading  {scores_path}")
    scores, run_meta = load_scores(scores_path)
    design = validate_design(scores)
    print(f"  validated: {design['n_specs']} specs x {design['n_conditions']} conditions "
          f"x {CHARTERS_PER_CELL} charters = {design['n_charters']}")

    matrices = build_matrices(scores, design["specs"])

    # IRR sources still live in the paper folder; look there if the package lacks them.
    irr_key = next((p for p in [repo / "data" / "irr-r1-key.json",
                                paper / "research" / "irr-r1-key.json"] if p.exists()), None)
    irr_xlsx = next((p for p in [repo / "data" / "Paper 2 IRR Scoring Sheet.xlsx",
                                 paper / "research" / "Paper 2 IRR Scoring Sheet.xlsx"]
                     if p.exists()), None)

    irr = None
    paired = None
    if irr_key and irr_xlsx:
        paired = load_irr(irr_key, irr_xlsx)
        if paired:
            irr = analyse_irr(paired)
            print(f"  IRR: {len(paired)} charters, "
                  f"{irr['overall']['two_point_n']} two-point disagreements, "
                  f"conditions covered {irr['overall']['conditions_covered']}")
    if irr is None:
        print("  IRR: sources not found — RQ3 values will be omitted")

    # Round 2: second independent human rater, 75-charter packet.
    irr_r2_key = next((p for p in [repo / "data" / "irr-r2-key.json",
                                   paper / "research" / "irr-r2-key.json"] if p.exists()), None)
    irr_r2_csv = next((p for p in [repo / "data" / "irr-r2-rater2-response.csv",
                                   paper / "research" / "irr-r2-rater2-response.csv"]
                       if p.exists()), None)
    irr_r2 = None
    if irr_r2_key and irr_r2_csv:
        paired_r2 = load_irr_r2(irr_r2_key, irr_r2_csv)
        if paired_r2:
            irr_r2 = build_irr_r2_bundle(paired_r2, paired)
            allr = irr_r2["all"]["overall"]
            msg = (f"  IRR round 2: {allr['n_charters']} charters "
                   f"({', '.join(allr['conditions_covered'])}), "
                   f"exact {allr['exact_pct']:.1f}%, {allr['two_point_n']} two-point, "
                   f"AC2 {allr['ac2']:.3f}, rater 2 "
                   f"{allr['human_mean_pct'] - allr['auto_mean_pct']:+.1f}pp vs automated")
            if "human_human" in irr_r2:
                hh = irr_r2["human_human"]["overall"]
                msg += (f"; human-human n={hh['n_charters']} exact {hh['exact_pct']:.1f}% "
                        f"AC2 {hh['ac2']:.3f} alpha {hh['alpha_ordinal']:.3f}")
            print(msg)
    if irr_r2 is None:
        print("  IRR round 2: sources not found — round 2 values will be omitted")

    # The earlier (March 2026, pre-Intermediate) scoring run used only by scorer_stability.
    # Shipped inside the package as data/scorer-stability-prior.json so the scorer-stability
    # threat statistic reproduces from a download; falls back to the authoring tree. The
    # name deliberately avoids the etcg-scores-*.json pattern the judge loader globs.
    stability_prior = next((p for p in [repo / "data" / "scorer-stability-prior.json",
                                        paper / "research" / "etcg-scores.json"]
                            if p.exists()), None)
    stability = scorer_stability(scores, stability_prior) if stability_prior else None
    if stability:
        print(f"  scorer stability: {stability['charters_changed']}/{stability['n_compared']} "
              f"charters changed between runs, max condition mean shift "
              f"{stability['max_condition_mean_shift']:.2f}pp")

    prerepair = prerepair_headline(prerepair_path)
    if prerepair:
        _pf = prerepair["contrasts"]["full_framework"]
        _cur = paired_contrast(matrices["spec_pct"]["etcg"],
                               matrices["spec_pct"]["baseline"], "ETCG vs Baseline")
        print(f"  pre-repair headline: Baseline "
              f"{prerepair['conditions']['baseline']['mean']:.2f}%, ETCG-Baseline "
              f"{_pf['mean_diff']:+.2f}pp p={_pf['p_signed_rank']:.3f}  "
              f"(corrected: {_cur['mean_diff']:+.2f}pp p={_cur['p_signed_rank']:.3f})")

    mixed = mixed_effects(scores)
    if mixed and "error" in mixed:
        print(f"  mixed model: {mixed['error']}")
    elif mixed:
        print(f"  mixed model: {mixed['optimizer']}, n={mixed['n_obs']}, "
              f"groups={mixed['n_groups']}, ICC={mixed['icc']:.4f}")

    vocab = approach_vocabulary(repo / "data" / "etcg-results.json")
    if vocab:
        print(f"  approach vocabulary: {vocab['n_distinct']} distinct labels "
              f"({vocab['n_distinct_normalised']} normalised) across "
              f"{vocab['n_charters']} ETCG charters, "
              f"{vocab['n_singletons']} used once")
    else:
        print("  approach vocabulary: etcg-results.json not found — values omitted")

    worked = worked_example_verbosity(repo)
    if worked:
        print(f"  worked example ({worked['spec_id']}): "
              + ", ".join(f"{CONDITION_LABELS[c]} {worked[c]}w" for c in CONDITIONS))
    else:
        print("  worked example: generation outputs not found — values omitted")

    gencost = generation_cost(repo / "data" / "generation-instrumentation.json")
    if gencost:
        e = gencost["per_condition"]["etcg"]
        print(f"  generation cost: ETCG ${e['cost_per_spec']:.4f}/spec "
              f"({e['latency_mean']:.1f}s, {e['tok_out_mean']:.0f} out-tok); "
              f"instrumentation run total ${gencost['total_cost_usd']:.2f}")
    else:
        print("  generation cost: instrumentation file absent or incomplete — values omitted")

    judges = judge_convergence(scores, matrices, design["specs"], repo / "data")
    if judges and judges.get("judges"):
        for slug, j in sorted(judges["judges"].items()):
            rep = j["replication"]["full_framework"]
            print(f"  judge {slug}: {j['label']} — overall {j['overall_judge_mean']:.2f}% "
                  f"vs GPT-4o {j['overall_primary_mean']:.2f}% "
                  f"(gap {j['overall_leniency_gap']:+.2f}pp), "
                  f"exact {j['exact_pct']:.1f}%, AC2 {j['gwet_ac2']:.3f}; "
                  f"ETCG-Baseline {rep['judge']['mean_diff']:+.2f}pp "
                  f"p={rep['judge']['p_signed_rank']:.3f}"
                  f"{' [ordering preserved]' if j['ranking_matches_primary'] else ' [ORDERING CHANGED]'}")
    if judges and judges.get("skipped"):
        for note in judges["skipped"]:
            print(f"  judge skipped: {note}")
    if not judges:
        print("  cross-model judges: no rescoring files found — values omitted")

    multimodel = multimodel_generalization(repo / "data" / "multimodel", matrices,
                                           design["specs"])
    if multimodel and multimodel.get("models"):
        for slug, md in sorted(multimodel["models"].items()):
            rg, sch = md["contrasts"]["role_guidance"], md["contrasts"]["schema"]
            print(f"  generator {slug}: {md['short']} — "
                  f"B/I/E {md['condition']['baseline']['mean']:.1f}/"
                  f"{md['condition']['intermediate']['mean']:.1f}/"
                  f"{md['condition']['etcg']['mean']:.1f}%  "
                  f"role {rg['mean_diff']:+.2f}pp p={rg['p_signed_rank']:.3f}  "
                  f"schema {sch['mean_diff']:+.2f}pp p={sch['p_signed_rank']:.3f}"
                  f"{'' if md['ranking_matches_primary'] else '  [RANKING DIFFERS]'}")
        print(f"  multi-model roll-up: role framing positive in "
              f"{multimodel['role_positive_in']}/{multimodel['n_generators']}, "
              f"schema non-significant in {multimodel['schema_nonsignificant_in']}/"
              f"{multimodel['n_generators']}, baseline last in "
              f"{multimodel['baseline_last_in']}/{multimodel['n_generators']}")
    if multimodel and multimodel.get("skipped"):
        for note in multimodel["skipped"]:
            print(f"  generator skipped: {note}")
    if not multimodel:
        print("  multi-model generation: no score sets found — values omitted")

    reg = Registry()
    register_all(reg, matrices, scores, design, irr, mixed, stability, vocab, worked, judges,
                 irr_r2, gencost, multimodel, prerepair)
    print(f"  registered {len(reg.keys())} values")

    inputs = {"scoring": file_meta(scores_path, repo),
              "scoring_prerepair": file_meta(prerepair_path, repo)}
    if stability_prior and stability_prior.exists():
        inputs["scoring_stability_prior"] = file_meta(stability_prior, repo)
    if irr_key:
        inputs["irr_key"] = file_meta(irr_key, repo)
    if irr_xlsx:
        inputs["irr_human_scores"] = file_meta(irr_xlsx, repo)
    if irr_r2_key:
        inputs["irr_r2_key"] = file_meta(irr_r2_key, repo)
    if irr_r2_csv:
        inputs["irr_r2_human_scores"] = file_meta(irr_r2_csv, repo)
    gen_instr_path = repo / "data" / "generation-instrumentation.json"
    if gencost and gen_instr_path.exists():
        inputs["generation_instrumentation"] = file_meta(gen_instr_path, repo)
    if multimodel and multimodel.get("models"):
        for slug in sorted(multimodel["models"]):
            sp = repo / "data" / "multimodel" / slug / "scores.json"
            if sp.exists():
                inputs[f"multimodel_{slug}_scores"] = file_meta(sp, repo)

    payload = {
        "generated": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "generator": "scripts/analysis.py",
        "scoring_run_metadata": run_meta,
        "inputs": inputs,
        "design": {k: v for k, v in design.items() if k != "specs"},
        "unit_of_analysis": "specification (n=25); charters within a specification are clustered",
        "values": reg.as_dict(),
        "mixed_effects": mixed,
        "irr_detail": irr,
        "irr_r2_detail": irr_r2,
        "scorer_stability": stability,
        "judge_convergence": judges,
        "generation_cost": gencost,
        "multimodel_generalization": multimodel,
        "prerepair_headline": prerepair,
    }
    (results_dir / "results.json").write_text(
        json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"wrote    {results_dir / 'results.json'}")

    mapping = emit_macros(reg, results_dir / "macros.tex")
    print(f"wrote    {results_dir / 'macros.tex'} ({len(mapping)} macros)")

    tables = emit_tables(reg, results_dir)
    for t in tables:
        print(f"wrote    {t}")

    emit_traceability(reg, mapping, results_dir / "TRACEABILITY.md", inputs, tables, irr, irr_r2)
    print(f"wrote    {results_dir / 'TRACEABILITY.md'}")

    # Mirror the LaTeX artefacts into the manuscript tree.
    gen_dir = paper / "submission" / "generated"
    gen_dir.mkdir(parents=True, exist_ok=True)
    for src in [results_dir / "macros.tex", *tables]:
        (gen_dir / src.name).write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
    print(f"mirrored {len(tables) + 1} .tex files -> {gen_dir}")

    if not args.no_figures:
        try:
            import figures
            made = figures.render_all(matrices, reg, repo / "figures", paper / "figures")
            for f in made:
                print(f"wrote    {f}")
        except ImportError:
            print("  figures.py not importable — skipping figures")
        except Exception as exc:
            print(f"  figure generation failed: {type(exc).__name__}: {exc}", file=sys.stderr)

    if args.audit_tex:
        target = args.audit_tex.resolve()
        print(f"\naudit    {target}")
        findings = audit_tex(reg, target)
        if findings:
            print(f"  {len(findings)} hand-typed decimal literal(s) not produced by this script:")
            for f in findings:
                print(f"    {f}")
        else:
            print("  clean — every decimal literal traces to a registered value")

    return 0


if __name__ == "__main__":
    sys.exit(main())
